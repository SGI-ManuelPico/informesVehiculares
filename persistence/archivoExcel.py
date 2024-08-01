import pandas as pd
import openpyxl
import re
import xlrd
import os
from openpyxl import load_workbook
from datetime import datetime 
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from util.tratadoArchivos import TratadorArchivos
from util.conversores import ConversoresExcel
import locale

locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')


class FuncionalidadExcel:
    def __init__(self):
        super().__init__()

    def extraerUbicar(self, file1, file2): 
        try:
            # Cargar el archivo de Excel
            workbook = openpyxl.load_workbook(file1)
            workbook2 = openpyxl.load_workbook(file2)

            sheet = workbook.active
            sheet2 = workbook2.active

            # Extraer la información necesaria del reporte
            placa_completa = sheet.cell(row=1, column=2).value
            # Extraer la parte relevante de la placa
            placa = placa_completa.split()[1] + placa_completa.split()[2]
            fecha = sheet.cell(row=2, column=2).value.split()[0].replace('-', '/')
            km_recorridos = float(sheet.cell(row=4, column=2).value.replace(' Km', ''))
            dia_trabajado = 1 if km_recorridos > 0 else 0
            preoperacional = 1 if dia_trabajado == 1 else None
            numExcesos = sheet.cell(row=9, column=2).value

            # Contar número de desplazamientos
            num_desplazamientos = 0
            for i in range(5, sheet2.max_row + 1):
                if sheet2.cell(row=i, column=1).value == 'Movimiento':
                    num_desplazamientos += 1

            # Crear el diccionario con los datos extraídos
            datos_extraidos = {
                'placa': placa,
                'fecha': fecha,
                'km_recorridos': km_recorridos,
                'dia_trabajado': dia_trabajado,
                'preoperacional': preoperacional,
                'num_excesos': numExcesos,
                'num_desplazamientos': num_desplazamientos,
            }

            self.datosUbicar = [datos_extraidos]

            return self.datosUbicar
        except Exception as e:
            print('Archivos incorrectos o faltantes UBICAR')
            return []

    # Extraer los datos de los informes de Ituran.

    def extraerIturan(self, file1, file2):

        try:
            # Cargar el archivo csv
            itu = pd.read_csv(file1)[['NICK_NAME', 'TOTAL_TRIP_DISTANCE', 'TOTAL_NUMBER_OF_TRIPS']]
            itu2 = pd.read_csv(file2)
            
            fecha = pd.to_datetime(itu2.loc[0, 'EVENT_START_DAY_TIME']).strftime('%m/%d/%Y')
            print(fecha)
            
            # Cambiar el nombre de las columnas
            itu = itu.rename(columns={
                'NICK_NAME': 'placa',
                'TOTAL_TRIP_DISTANCE': 'km_recorridos',
                'TOTAL_NUMBER_OF_TRIPS': 'num_desplazamientos'
            })
            
            # Agregar columna 'fecha'
            itu['fecha'] = fecha
            
            # Agregar columnas 'dia_trabajado' y 'preoperacional'
            itu['dia_trabajado'] = itu['km_recorridos'].apply(lambda x: 1 if x > 0 else 0)
            itu['preoperacional'] = itu['dia_trabajado'].apply(lambda x: 1 if x == 1 else 0)
            
            # Calcular el número de excesos de velocidad y crear DataFrame para fusiones
            excesos = itu2[itu2['TOP_SPEED'] > 80].groupby('V_NICK_NAME').size().reset_index(name='num_excesos')
            excesos = excesos.rename(columns={'V_NICK_NAME': 'placa'})
            
            # Unir el DataFrame de excesos con el DataFrame itu
            itu = itu.merge(excesos, on='placa', how='left')
            
            # Reemplazar los valores NaN en la columna num_excesos por 0
            itu['num_excesos'] = itu['num_excesos'].fillna(0).astype(int)
            
            # Convertir el DataFrame filtrado a un diccionario sin incluir el índice
            self.datos_extraidos = itu.to_dict(orient='records')
            
            return self.datos_extraidos
        
        # Si no se pueden sacar los archivos de la plataforma por alguna razón:

        except Exception as e: 
            print('Archivos incorrectos o faltantes ituran')
            return []

    # Extraer los datos de los informes de MDVR.

    def extraerMDVR(self, file1, file2):  # file2 es el informe general, file2 es el informe de paradas (para determinar los desplazamientos)

        try:   # Cargar el archivo de Excel usando openpyxl
            workbook = openpyxl.load_workbook(file1)
            sheet = workbook.active

            workbook2 = openpyxl.load_workbook(file2)
            sheet2 = workbook2.active

            # Extraer la información necesaria del reporte
            placa_completa = sheet.cell(row=1, column=2).value  
            placa = placa_completa.replace('-', '')  
            fecha = sheet.cell(row=2, column=2).value.split()[0] 
            km_recorridos = float(sheet.cell(row=4, column=2).value.replace(' Km', '')) 
            dia_trabajado = 1 if km_recorridos > 0 else 0
            preoperacional = 1 if dia_trabajado == 1 else 0
            num_excesos = int(sheet.cell(row=9, column=2).value)  # A9 es (9, 1)

            # Contar número de desplazamientos
            num_desplazamientos = 0
            for i in range(1, sheet2.max_row + 1):
                if sheet2.cell(row=i, column=2).value == 'Movimiento':
                    num_desplazamientos += 1

            # Crear el diccionario con los datos extraídos
            datos_extraidos = {
                'placa': placa,
                'fecha': fecha,
                'km_recorridos': km_recorridos,
                'dia_trabajado': dia_trabajado,
                'preoperacional': preoperacional,
                'num_excesos': num_excesos,
                'num_desplazamientos': num_desplazamientos,
            }
            self.datosMdvr = [datos_extraidos]
            return self.datosMdvr
        except Exception as e:
            return []

        
    # Extraer los datos de los informes de Securitrac.

    def extraerSecuritrac(self, file_path):

        try:
        # Cargar el archivo de Excel usando pandas
            df = pd.read_excel(file_path)

            # Diccionario para almacenar los datos por placa
            datos_por_placa = {}

            # Iterar sobre las filas del DataFrame
            for index, row in df.iterrows():
                placa = row['NROMOVIL']
                evento = row['EVENTO']
                kilometros = float(row['KILOMETROS'])
                fecha = row['FECHAGPS']

                # Formatear la fecha en dd/mm/aaaa
                fecha_formateada = pd.to_datetime(fecha).strftime('%d/%m/%Y')

                if placa not in datos_por_placa:
                    datos_por_placa[placa] = {
                        'placa': placa,
                        'fecha': fecha_formateada,  # Usar la fecha formateada
                        'km_recorridos': 0,
                        'num_excesos': 0,
                        'num_desplazamientos': 0
                    }
                datos_por_placa[placa]['km_recorridos'] += kilometros
                if evento == 'Exc. Velocidad':
                    datos_por_placa[placa]['num_excesos'] += 1
                if evento == 'Apagado':
                    datos_por_placa[placa]['num_desplazamientos'] += 1

            for placa, datos in datos_por_placa.items():
                datos['dia_trabajado'] = 1 if datos['km_recorridos'] > 0 else 0
                datos['preoperacional'] = 1 if datos['dia_trabajado'] == 1 else 0

            self.datos = []
            for x in datos_por_placa.keys():
                self.datos.append(datos_por_placa[x])

            return self.datos
        
        except Exception as e:
            print('Archivos incorrectos o faltantes SECURITRAC')
            return []

    # Extraer los datos de los informes de Ubicom.

    def extraerUbicom(self, file1, file2):

        try: 
            # Cargar el archivo de Excel usando xlrd
            workbook = xlrd.open_workbook(file1)
            sheet = workbook.sheet_by_index(0)
            workbook2 = xlrd.open_workbook(file2)
            sheet2 = workbook2.sheet_by_index(0)

            # Extraer la información necesaria del reporte
            fecha = sheet.cell_value(11, 11).split()[0]  # Celda L12
            
            km_recorridos = int(sheet.cell_value(20, 12))  # Celda M20
        
            num_excesos = int(sheet.cell_value(20, 21))  # Celda V20
        
            placa = sheet.cell_value(13, 24).split(' - ')[1].replace('(', '').replace(')', '') # Celda Y14, quitando el texto adicional
            

            # Calcular día trabajado y preoperacional
            dia_trabajado = 1 if km_recorridos > 0 else 0
            preoperacional = 1 if dia_trabajado == 1 else 0

            # Extraer el número de desplazamientos

        
            num_desplazamientos = 0
            for i in range(17, sheet2.nrows - 1):  # Empezamos en la fila 18 y ajustamos para no contar la última fila
                if sheet2.cell_value(i, 10) != '':  # Columna K es el indice 10
                    num_desplazamientos += 1

            # Crear el diccionario con los datos extraídos
            datos_extraidos = {
                'placa': placa,
                'fecha': fecha,
                'km_recorridos': km_recorridos,
                'dia_trabajado': dia_trabajado,
                'preoperacional': preoperacional,
                'num_excesos': num_excesos,
                'num_desplazamientos': num_desplazamientos,
            }

            self.datosUbicom = [datos_extraidos]

            return self.datosUbicom
        except Exception as e:
            print('Archivos incorrectos o faltantes UBICOM')
            return []

    # Extraer los datos de los informes de Wialon.

    def extraerWialon(self, file_path1, file_path2, file_path3):

        try:

            self.datos_extraidos = []
            
            for file_path in [file_path1, file_path2, file_path3]:
                xl = pd.ExcelFile(file_path)
                
                # Extraer placa y fecha siempre
                if 'Statistics' in xl.sheet_names:
                    statistics_df = xl.parse('Statistics', header=None)
                    placa = statistics_df.iloc[0, 1]  # Celda B1
                    fecha = statistics_df.iloc[1, 1]
                    
                    if isinstance(fecha, str):
                        fecha = pd.to_datetime(fecha.split()[0].replace('-', '/'))

                    fecha_formateada = fecha.strftime('%d/%m/%Y')
                    km_recorridos = int(statistics_df.iloc[7, 1])  # Celda B8, quitando 'km'
                    dia_trabajado = 1 if km_recorridos > 0 else 0
                    preoperacional = 1 if dia_trabajado == 1 else 0
                
                    datos = {
                        'placa': placa,
                        'fecha': fecha_formateada,
                        'km_recorridos': km_recorridos,
                        'dia_trabajado': dia_trabajado,
                        'preoperacional': preoperacional
                    }

                else:
                    datos = {
                        'placa': placa,
                        'fecha': fecha_formateada,
                        'km_recorridos': 0,
                        'dia_trabajado': 0,
                        'preoperacional': 0,
                    }
                
                # Verificar si el archivo tiene datos
                if 'Excesos de velocidad' in xl.sheet_names:
                    excesos_df = xl.parse('Excesos de velocidad', header=None)
                    num_excesos = len(excesos_df) - 1  # Descontar la fila de encabezado
                    datos.update({'num_excesos': num_excesos})
                    
                else:
                    datos.update({'num_excesos': 0})

                
                # Extraer número de desplazamientos
                if 'Cronología' in xl.sheet_names:
                    crono = xl.parse('Cronología')
                    desplazamientos = 0
                    for x in crono['Tipo'].to_list():
                        if x == 'Trip':
                            desplazamientos += 1
                    
                    datos.update({'num_desplazamientos': desplazamientos})

                else:
                    datos.update({'num_desplazamientos': 0})
                    
                
                self.datos_extraidos.append(datos)
            
            return self.datos_extraidos
        
        except Exception as e: 
            print('Archivos incorrectos o faltantes WIALON')
            return []

    # Infractores diario Ubicar

    def infracUbicar(self, file1):

        try:
            # Leer el archivo de Excel y obtener las hojas
            df = pd.read_excel(file1, skiprows=2).iloc[:-1]  # Ignorar las primeras 4 filas y la última fila

            # Extraer la placa del vehículo de la celda B1
            placa = "JYT620"  # Reemplazar con la extracción correcta si se requiere

            # Convertir la columna 'Comienzo' a datetime y extraer solo la fecha y hora
            df['Comienzo'] = pd.to_datetime(df['Comienzo'], dayfirst= True)
            df['Fecha'] = df['Comienzo'].dt.strftime('%d/%m/%Y %H:%M:%S')
            
            df['Tiempo de Exceso'] = df['Duración'].apply(ConversoresExcel().conversorSegundosUbicar())
            # Crear el diccionario con el formato requerido
            self.registros = []
            for index, row in df.iterrows():
                self.registros.append({
                    'PLACA': placa,
                    'TIEMPO DE EXCESO': row['Tiempo de Exceso'],
                    'DURACIÓN EN KM DE EXCESO': None,
                    'VELOCIDAD MÁXIMA': float(row['Velocidad máxima'].replace(' kph', '')),
                    'PROYECTO': None,
                    'CONDUCTOR': None,
                    'RUTA DE EXCESO': row['Posición'], # Por ahora guardamos las coordenadas porque pasarlas a dirección requiere de otras cosas.
                    'FECHA': row['Fecha']
                })

            # Imprimir el número de registros
            print(f'Número de registros: {len(self.registros)}')
            
            # Retornar el diccionario de registros
            return self.registros
        except Exception as e:

            return []

    # Infractores diario MDVR

    def infracMDVR(self, file):

        try:
            # Abrir el archivo Excel ignorando posibles corrupciones
            workbook = xlrd.open_workbook(file, ignore_workbook_corruption=True)
            
            # Leer el archivo Excel ignorando las primeras dos filas y la última fila
            df = pd.read_excel(workbook, skiprows=2).iloc[:-1]

            # Extraer la placa del vehículo de la celda B1
            placa = pd.read_excel(workbook).columns[1].replace(' ', '')

            # Convertir la columna 'Comienzo' a datetime y extraer la fecha y hora
            df['Comienzo'] = pd.to_datetime(df['Comienzo'], dayfirst=True)
            df['Fecha'] = df['Comienzo'].dt.strftime('%d/%m/%Y %H:%M:%S')

            # Dividir la columna 'Posición' en 'Latitud' y 'Longitud'
            df[['Latitud', 'Longitud']] = df['Posición'].str.split(',', expand=True)
            df['Latitud'] = df['Latitud'].astype(float)
            df['Longitud'] = df['Longitud'].astype(float)

            df['Duración exceso de velocidad'] = df['Duración exceso de velocidad'].apply(ConversoresExcel().conversorSegundosMDVR())

            # Crear el diccionario en el formato requerido
            self.registros = []
            for index, row in df.iterrows():
                registro = {
                    'PLACA': placa,
                    'TIEMPO DE EXCESO': row['Duración exceso de velocidad'],
                    'DURACIÓN EN KM DE EXCESO': None,
                    'VELOCIDAD MÁXIMA': float(row['Velocidad máxima'].replace('kph', '')),
                    'PROYECTO': '',
                    'RUTA DE EXCESO': f"{row['Latitud']}, {row['Longitud']}",  # Por ahora guardamos las coordenadas porque pasarlas a dirección requiere de otras cosas.
                    'CONDUCTOR': '',
                    'FECHA': row['Fecha']
                }
                self.registros.append(registro)

            print(f"Número total de registros: {len(self.registros)}")
            return self.registros
        
        except Exception as e:
            return []


    # Infractores diario Securitrac

    def infracSecuritrac(self, file):
        try:

            # Leer el archivo de Excel
            df = pd.read_excel(file)

            # Filtrar solo las filas con "Exc. Velocidad"
            df = df[df['EVENTO'] == 'Exc. Velocidad']

            # Crear el diccionario con el formato requerido
            self.registros = []
            for index, row in df.iterrows():
                fecha_formateada = pd.to_datetime(row['FECHAGPS']).strftime('%d/%m/%Y %H:%M:%S')
                registro = {
                    'PLACA': row['NROMOVIL'],
                    'TIEMPO DE EXCESO': None,
                    'DURACIÓN EN KM DE EXCESO': None,
                    'VELOCIDAD MÁXIMA': row['VELOCIDAD'],
                    'RUTA DE EXCESO': row['POSICION'], 
                    'PROYECTO': None,
                    'CONDUCTOR': None,
                    'FECHA': fecha_formateada,
                }
                self.registros.append(registro)

            print(f"Total de registros generados: {len(self.registros)}")
            print(self.registros[:5]) # Muestra los primeros 5 registros para verificar

            return self.registros

        except Exception as e:
            return []

    # Infractores diario Ituran

    def infracIturan(self, file1):

        try: 
            # Leer el archivo de Excel y obtener las hojas
            df = pd.read_csv(file1)
            
            # Filtrar la hoja que contiene la información relevante
            df_infracciones = df[['V_NICK_NAME', 'EVENT_DURATION_SEC', 'EVENT_DISTANCE', 'TOP_SPEED', 'VEHICLE_GROUP', 'ADDRESS', 'DRIVER_NAME', 'EVENT_START_DAY_TIME']]

            # Renombrar las columnas según lo solicitado
            df_infracciones.columns = ['PLACA', 'TIEMPO DE EXCESO', 'DURACIÓN EN KM DE EXCESO', 'VELOCIDAD MÁXIMA', 'PROYECTO', 'RUTA DE EXCESO', 'CONDUCTOR', 'FECHA']

            # Mantener la hora en la columna FECHA y convertirla al formato dd/mm/yyyy HH:MM:SS
            df_infracciones['FECHA'] = pd.to_datetime(df_infracciones['FECHA']).dt.strftime('%d/%m/%Y %H:%M:%S')

            # Filtrar los registros relevantes
            df_infracciones = df_infracciones[df_infracciones['DURACIÓN EN KM DE EXCESO'] > 0]

            num_registros = len(df_infracciones)
            print(f"Número de registros: {num_registros}")

            # Convertir el DataFrame en un diccionario
            self.datos_infracciones = df_infracciones.to_dict(orient='records')

            return self.datos_infracciones
        
        except Exception as e:
            return []

    # Infractores diario Wialon

    def infracWialon(self, file1):

        try:
            xls = pd.ExcelFile(file1)

            # Verificar si la hoja 'Excesos de velocidad' existe
            if 'Excesos de velocidad' not in xls.sheet_names:
                print("La hoja 'Excesos de velocidad' no está presente en el archivo.")
                return []

            # Leer la hoja 'Statistics' para obtener la placa
            df_stats = pd.read_excel(xls, 'Statistics')
            placa = df_stats.columns[1]  # Celda B1 en la hoja "Statistics"

            

            # Leer la hoja 'Excesos de velocidad' para obtener la información necesaria
            df_excesos = pd.read_excel(xls, 'Excesos de velocidad')

            # Convertir la columna 'Comienzo' a datetime y extraer solo la fecha y hora
            df_excesos['Comienzo'] = pd.to_datetime(df_excesos['Comienzo'])
            df_excesos['Fecha'] = df_excesos['Comienzo'].dt.strftime('%d/%m/%Y %H:%M:%S')

            df_excesos['Duración en segundos'] = df_excesos['Duración'].apply(ConversoresExcel().conversorSegundosWialon())

            # Crear el diccionario con el formato requerido
            self.registros = []
            for index, row in df_excesos.iterrows():
                registro = {
                    'PLACA': placa,
                    'TIEMPO DE EXCESO': row['Duración en segundos'],
                    'DURACIÓN EN KM DE EXCESO': '',  # No disponible en este documento
                    'VELOCIDAD MÁXIMA': row['Velocidad máxima'],
                    'PROYECTO': '',  # No disponible en este documento
                    'CONDUCTOR': '',  # No disponible en este documento
                    'RUTA DE EXCESO': row['Localización'],
                    'FECHA': row['Fecha']
                }
                self.registros.append(registro)

            print(f"Total de registros extraídos: {len(self.registros)}")
            return self.registros# Extraer información infractores Wialon
        
        except Exception as e:
            return []


    # Odómetro Ituran

    def OdomIturan(self, file):
        # Leer el archivo de Excel
        od = pd.read_csv(file)

        # Extraer la placa y el odómetro

        df = od[['V_PLATE_NUMBER', 'END_ODOMETER']]

        # Renombrar las columnas

        df.columns = ['PLACA', 'KILOMETRAJE']

        # Crear el diccionario con el formato requerido

        self.datos = df.to_dict('records')
        return self.datos

    # Odómetro Ubicar 

    def odomUbicar(self, file):
        # Leer el archivo de Excel
        df = pd.read_excel(file)  

        # Extraer la placa del vehículo de la celda B1, si agregan otro carro a esta plataforma toca cambiar como se extrae esto.
        placa = 'JYT620'

        # Extraer el odómetro de la celda correspondiente
        odometro = df.iloc[11, 2] 

        # Crear el diccionario con el formato requerido
        registro = {
            'PLACA': placa,
            'KILOMETRAJE': float(odometro.split()[0].replace(',', ''))
        }

        self.registroUbicar = [registro]
        return self.registroUbicar


    ## ACTUALIZAR HOJAS DE INDICADORES

    # Este es el DF que se usa para actualiza la hoja de Indicadores. Necesita que se haya guardado ya df_exist, que se genera al correr crear_excel. Por esta razón, toca guardar en una variable df_exist, lo que retorna crear_excel. 
    # El df que retorna la función dfDiario también toca guardarlo como una variable, pues este es el que se usa para calcular todos los indicadores. 

    #EJL 

    def calcular_EJL(self, df_diario):

        # Crear una columna para el mes y año
        df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')
        
        # Calcular los acumulados mensuales
        df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()
        
        # Crear un DataFrame con todos los meses del año
        all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
        df_all_months = pd.DataFrame(all_months, columns=['MES'])
        
        # Unir con el DataFrame de acumulados mensuales para asegurarse de que todos los meses estén presentes
        df_acumulados = pd.merge(df_all_months, df_acumulados, on='MES', how='left')
        
        # Rellenar los NaN con ceros y asegurarse de que el tipo de datos sea correcto
        df_acumulados.fillna(0, inplace=True)
        
        # Cambiar el formato del periodo a nombre de mes en español
        df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()
        
        # Calcular EJL
        self.df_EJL = pd.DataFrame({
            'MES': df_acumulados['MES'],
            'EDD': df_acumulados['EXCESOS VELOCIDAD'],
            'SDT': df_acumulados['DÍA TRABAJADO'],
            'EJL': (df_acumulados['EXCESOS VELOCIDAD'] / df_acumulados['DÍA TRABAJADO']) * 100
        })
        
        # Redondear a 2 decimales
        self.df_EJL = self.df_EJL.round(2)
        
        # Transponer el DataFrame
        self.df_EJL = self.df_EJL.set_index('MES').transpose()
        
        return self.df_EJL

    ## GVE

    def calcular_GVE(self, df_diario, df_exist):

        # Crear una columna para el mes y año
        df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')
        
        # Calcular los acumulados mensuales
        df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()
        
        # Calcular el valor máximo de "DÍA TRABAJADO" por mes
        df_max_dia_trabajado = df_diario.groupby('MES')['DÍA TRABAJADO'].max().reset_index()
        
        # Crear un DataFrame con todos los meses del año
        all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
        
        # Unir los DataFrames de acumulados mensuales y máximos de día trabajado para asegurarse de que todos los meses estén presentes
        df_acumulados = pd.merge(all_months.to_frame(name='MES'), df_acumulados, on='MES', how='left')
        df_max_dia_trabajado = pd.merge(all_months.to_frame(name='MES'), df_max_dia_trabajado, on='MES', how='left')
        
        # Rellenar los NaN con ceros
        df_acumulados.fillna(0, inplace=True)
        df_max_dia_trabajado.fillna(0, inplace=True)

        # Cambiar el formato de la fecha para que muestre el nombre del mes.
        df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()
        df_max_dia_trabajado['MES'] = df_max_dia_trabajado['MES'].dt.strftime('%B').str.capitalize()
        
        # Crear un diccionario para mapear columnas a meses
        month_dict = {}
        for idx, col in enumerate(df_exist.columns):
            if '/' in col:
                month = col.split('/')[1]
                if month not in month_dict:
                    month_dict[month] = []
                month_dict[month].append(col)
        
        # Sumar valores para cada mes
        month_sums = pd.DataFrame()
        for month, cols in month_dict.items():
            month_sums[month] = df_exist[cols].sum(axis=1)
        
        # Combinar las columnas de identificación originales con las sumas mensuales
        result = pd.concat([df_exist.iloc[:, :2], month_sums], axis=1)
        
        # Calcular VIP para cada mes (cantidad de placas con al menos un 'Nº Desplazamiento' > 0)
        vip_values = []
        for month in all_months.strftime('%m'):
            vip_count = sum((result['SEGUIMIENTO'] == 'Nº Desplazamiento') & (result[month] > 0))
            vip_values.append(vip_count)
        
        # Calcular VLD (el valor máximo de DÍA TRABAJADO para cada mes)
        vld_values = df_max_dia_trabajado['DÍA TRABAJADO'].tolist()
        
        # Calcular GVE
        gve_values = [(vip / vld) * 100 if vld != 0 else 0 for vip, vld in zip(vip_values, vld_values)]
        
        # Crear el DataFrame de GVE
        self.df_GVE = pd.DataFrame({
            'MES': df_acumulados['MES'],
            'VIP': vip_values,
            'VLD': vld_values,
            'GVE': gve_values
        })
        
        # Redondear a 2 decimales
        self.df_GVE = self.df_GVE.round(2)
        
        # Transponer el DataFrame
        self.df_GVE = self.df_GVE.set_index('MES').transpose()
        
        return self.df_GVE

    # ELVL

    def calcular_ELVL(self, df_diario):

        # Crear una columna para el mes y año
        df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')

        # Calcular los acumulados mensuales
        df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()

        # Crear un DataFrame con todos los meses del año
        all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
        df_all_months = pd.DataFrame(all_months, columns=['MES'])

        # Unir con el DataFrame de acumulados mensuales para asegurarse de que todos los meses estén presentes
        df_acumulados = pd.merge(df_all_months, df_acumulados, on='MES', how='left')

        # Rellenar los NaN con ceros y asegurarse de que el tipo de datos sea correcto
        df_acumulados.fillna(0, inplace=True)

        # Cambiar el formato del periodo a nombre de mes en español
        df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()

        # Calcular ELVL
        self.df_ELVL = pd.DataFrame({
            'MES': df_acumulados['MES'],
            'DLEV': df_acumulados['EXCESOS VELOCIDAD'],
            'TDL': df_acumulados['DESPLAZAMIENTOS'],
            'ELVL': (df_acumulados['EXCESOS VELOCIDAD'] / df_acumulados['DESPLAZAMIENTOS']) * 100
        })

        # Redondear a 2 decimales
        self.df_ELVL = self.df_ELVL.round(2)

        # Transponer el DataFrame
        self.df_ELVL = self.df_ELVL.set_index('MES').transpose()

        return self.df_ELVL

    # IDP

    def calcular_IDP(self, df_diario):


        # Crear una columna para el mes y año
        df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')

        # Calcular los acumulados mensuales
        df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()

        # Crear un DataFrame con todos los meses del año
        all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
        df_all_months = pd.DataFrame(all_months, columns=['MES'])

        # Unir con el DataFrame de acumulados mensuales para asegurarse de que todos los meses estén presentes
        df_acumulados = pd.merge(df_all_months, df_acumulados, on='MES', how='left')

        # Rellenar los NaN con ceros
        df_acumulados.fillna(0, inplace=True)

        # Cambiar el formato de los meses.
        df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()

        # Calcular IDP
        self.df_IDP = pd.DataFrame({
            'MES': df_acumulados['MES'],
            'DPV': df_acumulados['PREOPERACIONAL'],
            'DTR': df_acumulados['DÍA TRABAJADO'],
            'IDP': (df_acumulados['PREOPERACIONAL'] / df_acumulados['DÍA TRABAJADO']) * 100
        })

        # Redondear a 2 decimales
        self.df_IDP = self.df_IDP.round(2)

        # Transponer el DataFrame
        self.df_IDP = self.df_IDP.set_index('MES').transpose()

        return self.df_IDP
    

    # Fuera de horario laboral

    def fueraLaboralSecuritrac(self, file_path):
    # Cargar el archivo Excel

        try:
            secudf = pd.read_excel(file_path)

            # Convertir 'FECHAGPS' a datetime
            secudf['FECHAGPS'] = pd.to_datetime(secudf['FECHAGPS'])

            # Filtrar filas donde 'EVENTO' es 'Encendido' y la hora es antes de las 6 AM o después de las 6 PM
            filtered_df = secudf[(secudf['EVENTO'] == 'Encendido') & ((secudf['FECHAGPS'].dt.hour < 6) | (secudf['FECHAGPS'].dt.hour >= 18))]

            # Verificar si el DataFrame filtrado está vacío
            if filtered_df.empty:
                return []

            # Seleccionar solo las columnas requeridas y renombrarlas
            filtered_df = filtered_df[['NROMOVIL', 'FECHAGPS']].rename(columns={'NROMOVIL': 'placa', 'FECHAGPS': 'fecha'})

            # Formatear la columna 'fecha'
            filtered_df['fecha'] = filtered_df['fecha'].dt.strftime('%d/%m/%Y %H:%M')

            # Convertir el dataframe filtrado a un diccionario
            self.result_dict = filtered_df.to_dict(orient='records')

            # Devolver el resultado dentro de una lista
            return self.result_dict
        except Exception as e:
            return []
    
    def fueraLaboralIturan(self, file_path):
        try:
            # Cargar el archivo CSV
            df = pd.read_csv(file_path)

            # Convertir 'TRIP_START_TIME' y 'TRIP_END_TIME' a datetime
            df['TRIP_START_TIME'] = pd.to_datetime(df['TRIP_START_TIME'])
            df['TRIP_END_TIME'] = pd.to_datetime(df['TRIP_END_TIME'])

            # Filtrar filas donde la hora de 'TRIP_START_TIME' o 'TRIP_END_TIME' es antes de las 6 AM o después de las 6 PM
            filtered_df = df[(df['TRIP_START_TIME'].dt.hour < 6) | (df['TRIP_START_TIME'].dt.hour >= 18) | (df['TRIP_END_TIME'].dt.hour < 6) | (df['TRIP_END_TIME'].dt.hour >= 18)]

            # Verificar si el DataFrame filtrado está vacío
            if filtered_df.empty:
                return []

            # Crear una nueva columna 'fecha' que tome el valor correcto basado en la condición
            filtered_df['fecha'] = filtered_df.apply(lambda row: row['TRIP_END_TIME'] if row['TRIP_END_TIME'].hour < 6 or row['TRIP_END_TIME'].hour >= 18 else row['TRIP_START_TIME'], axis=1)

            # Seleccionar solo las columnas requeridas y renombrarlas
            filtered_df = filtered_df[['V_NICK_NAME', 'fecha']].rename(columns={'V_NICK_NAME': 'placa'})

            # Formatear la columna 'fecha'
            filtered_df['fecha'] = filtered_df['fecha'].dt.strftime('%d/%m/%Y %H:%M')

            # Convertir el dataframe filtrado a un diccionario
            self.result_dict = filtered_df.to_dict(orient='records')

            # Devolver el resultado dentro de una lista
            return self.result_dict
        except Exception as e:
            return []
    
    def fueraLaboralMDVR(self, file_path):
        try:
            # Cargar el archivo Excel
            md = pd.read_excel(file_path, header=2, skipfooter=8)
            
            # Eliminar la primera fila
            md = md.iloc[1:]

            # Convertir 'Comienzo' y 'Fin' a datetime
            md['Comienzo'] = pd.to_datetime(md['Comienzo'])
            md['Fin'] = pd.to_datetime(md['Fin'])

            # Filtrar filas donde 'Estado' es 'Movimiento'
            md = md[md['Estado'] == 'Movimiento']

            # Filtrar filas donde la hora de 'Comienzo' o 'Fin' es antes de las 6 AM o después de las 6 PM
            filtered_df = md[(md['Comienzo'].dt.hour < 6) | (md['Comienzo'].dt.hour >= 18) | (md['Fin'].dt.hour < 6) | (md['Fin'].dt.hour >= 18)]

            # Verificar si el DataFrame filtrado está vacío
            if filtered_df.empty:
                return []

            # Crear una nueva columna 'fecha' que tome el valor correcto basado en la condición
            filtered_df['fecha'] = filtered_df.apply(lambda row: row['Fin'] if row['Fin'].hour < 6 or row['Fin'].hour >= 18 else row['Comienzo'], axis=1)

            # Seleccionar solo las columnas requeridas y renombrarlas
            filtered_df = filtered_df[['Vehiculo', 'fecha']].rename(columns={'Vehiculo': 'placa'})

            # Formatear la columna 'fecha'
            filtered_df['fecha'] = filtered_df['fecha'].dt.strftime('%d/%m/%Y %H:%M')

            # Convertir el dataframe filtrado a un diccionario
            self.result_dict = filtered_df.to_dict(orient='records')

            # Devolver el resultado dentro de una lista
            return self.result_dict
        except Exception as e:
            return []
    
    def fueraLaboralUbicar(self, file_path):
        try:
            # Cargar el archivo Excel
            ubi = pd.read_excel(file_path, header=2, skipfooter=11)
            
            # Eliminar la primera fila
            ubi = ubi.iloc[1:]

            # Convertir 'Comienzo' y 'Fin' a datetime
            ubi['Comienzo'] = pd.to_datetime(ubi['Comienzo'], dayfirst=True)
            ubi['Fin'] = pd.to_datetime(ubi['Fin'], dayfirst=True)

            # Filtrar filas donde 'Estado' es 'Movimiento'
            ubi = ubi[ubi['Estado'] == 'Movimiento']

            # Filtrar filas donde la hora de 'Comienzo' o 'Fin' es antes de las 6 AM o después de las 6 PM
            filtered_df = ubi[(ubi['Comienzo'].dt.hour < 6) | (ubi['Comienzo'].dt.hour >= 18) | (ubi['Fin'].dt.hour < 6) | (ubi['Fin'].dt.hour >= 18)]

            # Verificar si el DataFrame filtrado está vacío
            if filtered_df.empty:
                return []

            # Crear una nueva columna 'fecha' que tome el valor correcto basado en la condición
            fecha = filtered_df.apply(lambda row: row['Fin'] if row['Fin'].hour < 6 or row['Fin'].hour >= 18 else row['Comienzo'], axis=1)
            
            # Añadir la columna 'fecha' al DataFrame filtrado
            filtered_df = filtered_df.assign(fecha=fecha)

            # Añadir una columna constante 'placa'
            filtered_df['placa'] = 'JYT620'

            # Seleccionar solo las columnas requeridas
            filtered_df = filtered_df[['placa', 'fecha']]

            # Formatear la columna 'fecha'
            filtered_df['fecha'] = filtered_df['fecha'].dt.strftime('%d/%m/%Y %H:%M')

            # Convertir el dataframe filtrado a un diccionario
            self.result_dict = filtered_df.to_dict(orient='records')

            # Devolver el resultado dentro de una lista
            return self.result_dict
        except Exception as e:
            return []
    
    def fueraLaboralWialon(self, file_path):
        try:
            # Cargar el archivo Excel
            xl = pd.ExcelFile(file_path)
            
            # Extraer la placa
            placa = xl.parse(sheet_name='Statistics').columns[1]
            print(placa)
            if 'Cronología' in xl.sheet_names:
                # Cargar la hoja 'Cronología'
                df = xl.parse(sheet_name='Cronología')
                
                # Filtrar filas donde 'Tipo' es 'Trip'
                df = df[df['Tipo'] == 'Trip']

                # Convertir 'Comienzo' y 'Fin' a datetime
                df['Comienzo'] = pd.to_datetime(df['Comienzo'])
                df['Fin'] = pd.to_datetime(df['Fin'])

                # Filtrar filas donde la hora de 'Comienzo' o 'Fin' es antes de las 6 AM o después de las 6 PM
                filtered_df = df[(df['Comienzo'].dt.hour < 6) | (df['Comienzo'].dt.hour >= 18) | (df['Fin'].dt.hour < 6) | (df['Fin'].dt.hour >= 18)]

                # Verificar si el DataFrame filtrado está vacío
                if filtered_df.empty:
                    return []

                # Crear una nueva columna 'fecha' que tome el valor correcto basado en la condición
                fecha = filtered_df.apply(lambda row: row['Fin'] if row['Fin'].hour < 6 or row['Fin'].hour >= 18 else row['Comienzo'], axis=1)
                
                # Añadir la columna 'fecha' al DataFrame filtrado
                filtered_df = filtered_df.assign(fecha=fecha)

                # Añadir una columna constante 'placa'
                filtered_df['placa'] = placa

                # Seleccionar solo las columnas requeridas
                filtered_df = filtered_df[['placa', 'fecha']]

                # Formatear la columna 'fecha'
                filtered_df['fecha'] = filtered_df['fecha'].dt.strftime('%d/%m/%Y %H:%M')

                # Convertir el dataframe filtrado a un diccionario
                self.result_dict = filtered_df.to_dict(orient='records')

                # Devolver el resultado dentro de una lista
                return self.result_dict
            
            else: 
                return [] 
        except Exception as e:
            return []    


    def fueraLaboralTodos(self, rutasLaboral):
    
        all_results = []
        all_results.extend(self.fueraLaboralSecuritrac(rutasLaboral['securitrac']))
        all_results.extend(self.fueraLaboralMDVR(rutasLaboral['mdvr']))
        all_results.extend(self.fueraLaboralUbicar(rutasLaboral['ubicar']))
        all_results.extend(self.fueraLaboralIturan(rutasLaboral['ituran']))
    
        for file_path in rutasLaboral['wialon']:
            all_results.extend(self.fueraLaboralWialon(file_path))
        
        self.todosDF = pd.DataFrame(all_results)
        self.todosDF['fecha'] = pd.to_datetime(self.todosDF['fecha'], format='%d/%m/%Y %H:%M').dt.strftime('%Y-%m-%d %H:%M:%S')

        return self.todosDF  
