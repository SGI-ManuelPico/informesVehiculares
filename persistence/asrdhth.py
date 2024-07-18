import pandas as pd
import openpyxl
import xlrd
import os
from datetime import datetime 
import locale
import glob

locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
print("AAAAAAA")
# Extraer información del informe de Ubicar.

def xlsx(input_file):
   
    base_name = os.path.splitext(input_file)[0]
    output_file = base_name + '.xlsx'
    df = pd.read_excel(input_file, engine='xlrd')
    df.to_excel(output_file, index=False)
    return output_file
print("AAAAAAA")
def extraerUbicar(file1, file2): 
    try:
        # Cargar el archivo de Excel
        workbook = openpyxl.load_workbook(file1)
        workbook2 = openpyxl.load_workbook(file2)

        sheet = workbook.active
        sheet2 = workbook2.active

        # Extraer la información necesaria del reporte
        placa_completa = sheet.cell(row=1, column=2).value
        # Extraer la parte relevante de la placa
        placa = placa_completa.split()[1] + placa_completa.split()[2]
        fecha = sheet.cell(row=2, column=2).value.split()[0].replace('-', '/')
        km_recorridos = float(sheet.cell(row=4, column=2).value.replace(' Km', ''))
        dia_trabajado = 1 if km_recorridos > 0 else 0
        preoperacional = 1 if dia_trabajado == 1 else None
        numExcesos = sheet.cell(row=9, column=2).value

        # Contar número de desplazamientos
        num_desplazamientos = 0
        for i in range(5, sheet2.max_row + 1):
            if sheet2.cell(row=i, column=1).value == 'Movimiento':
                num_desplazamientos += 1

        # Crear el diccionario con los datos extraídos
        datos_extraidos = {
            'placa': placa,
            'fecha': fecha,
            'km_recorridos': km_recorridos,
            'dia_trabajado': dia_trabajado,
            'preoperacional': preoperacional,
            'num_excesos': numExcesos,
            'num_desplazamientos': num_desplazamientos,
        }

        return [datos_extraidos]
    except Exception as e:
        print('Archivos incorrectos o faltantes UBICAR')
        return []
print("AAAAAAA")
# Extraer los datos de los informes de Ituran.

def extraerIturan(file1, file2):

    try:
        # Cargar el archivo csv
        itu = pd.read_csv(file1)[['NICK_NAME', 'TOTAL_TRIP_DISTANCE', 'TOTAL_NUMBER_OF_TRIPS']]
        itu2 = pd.read_csv(file2)
        
        fecha = datetime.now().strftime('%d/%m/%Y')
        
        # Cambiar el nombre de las columnas
        itu = itu.rename(columns={
            'NICK_NAME': 'placa',
            'TOTAL_TRIP_DISTANCE': 'km_recorridos',
            'TOTAL_NUMBER_OF_TRIPS': 'num_desplazamientos'
        })
        
        # Agregar columna 'fecha'
        itu['fecha'] = fecha
        
        # Agregar columnas 'dia_trabajado' y 'preoperacional'
        itu['dia_trabajado'] = itu['km_recorridos'].apply(lambda x: 1 if x > 0 else 0)
        itu['preoperacional'] = itu['dia_trabajado'].apply(lambda x: 1 if x == 1 else 0)
        
        # Calcular el número de excesos de velocidad y crear DataFrame para fusiones
        excesos = itu2[itu2['TOP_SPEED'] > 80].groupby('V_NICK_NAME').size().reset_index(name='num_excesos')
        excesos = excesos.rename(columns={'V_NICK_NAME': 'placa'})
        
        # Unir el DataFrame de excesos con el DataFrame itu
        itu = itu.merge(excesos, on='placa', how='left')
        
        # Reemplazar los valores NaN en la columna num_excesos por 0
        itu['num_excesos'] = itu['num_excesos'].fillna(0).astype(int)
        
        # Convertir el DataFrame filtrado a un diccionario sin incluir el índice
        datos_extraidos = itu.to_dict(orient='records')
        
        return datos_extraidos
    
    # Si no se pueden sacar los archivos de la plataforma por alguna razón:

    except Exception as e: 
        print('Archivos incorrectos o faltantes ituran')
        return []
print("AAAAAAA")
# Extraer los datos de los informes de MDVR.

def extraerMDVR(file1, file2): #file1 es el informe general, file2 es el informe de paradas (para determinar los desplazamientos)

    file2x = xlsx(file2)

    try:
    # Cargar el archivo de Excel usando xlrd
        workbook = xlrd.open_workbook(file1)
        sheet = workbook.sheet_by_index(0)
        workbook2 = openpyxl.load_workbook(file2x)
        sheet2 = workbook2.active


        # Extraer la información necesaria del reporte
        placa_completa = sheet.cell_value(0, 1)  # A1 es (0, 1)
        placa = placa_completa.replace('-', '')  # Quitar el guion de la placa
        fecha = sheet.cell_value(1, 1).split()[0]  # A2 es (1, 1)
        km_recorridos = float(sheet.cell_value(3, 1).replace(' Km', ''))  # A5 es (4, 1)
        dia_trabajado = 1 if km_recorridos > 0 else 0
        preoperacional = 1 if dia_trabajado == 1 else 0
        num_excesos = int(sheet.cell_value(8, 1))  # A9 es (8, 1)

        #Contar número de desplazamientos
        num_desplazamientos = 0
        for i in range(1, sheet2.max_row + 1):
            if sheet2.cell(row=i, column=2).value == 'Movimiento':
                num_desplazamientos += 1

        # Crear el diccionario con los datos extraídos
        datos_extraidos = {
            'placa': placa,
            'fecha': fecha,
            'km_recorridos': km_recorridos,
            'dia_trabajado': dia_trabajado,
            'preoperacional': preoperacional,
            'num_excesos': num_excesos,
            'num_desplazamientos': num_desplazamientos,
        }

        return [datos_extraidos]
    except Exception as e: 
        return [] 
print("AAAAAAA")
# Extraer los datos de los informes de Securitrac.

def extraerSecuritrac(file_path):

    try:
    # Cargar el archivo de Excel usando pandas
        df = pd.read_excel(file_path)

        # Diccionario para almacenar los datos por placa
        datos_por_placa = {}

        # Iterar sobre las filas del DataFrame
        for index, row in df.iterrows():
            placa = row['NROMOVIL']
            evento = row['EVENTO']
            kilometros = float(row['KILOMETROS'])
            fecha = row['FECHAGPS']

            # Formatear la fecha en dd/mm/aaaa
            fecha_formateada = pd.to_datetime(fecha).strftime('%d/%m/%Y')

            if placa not in datos_por_placa:
                datos_por_placa[placa] = {
                    'placa': placa,
                    'fecha': fecha_formateada,  # Usar la fecha formateada
                    'km_recorridos': 0,
                    'num_excesos': 0,
                    'num_desplazamientos': 0
                }
            datos_por_placa[placa]['km_recorridos'] += kilometros
            if evento == 'Exc. Velocidad':
                datos_por_placa[placa]['num_excesos'] += 1
            datos_por_placa[placa]['num_desplazamientos'] += 1

        for placa, datos in datos_por_placa.items():
            datos['dia_trabajado'] = 1 if datos['km_recorridos'] > 0 else 0
            datos['preoperacional'] = 1 if datos['dia_trabajado'] == 1 else 0

        datos = []
        for x in datos_por_placa.keys():
            datos.append(datos_por_placa[x])

        return datos
    
    except Exception as e:
        print('Archivos incorrectos o faltantes SECURITRAC')
        return []
print("AAAAAAA")
# Extraer los datos de los informes de Ubicom.

def extraerUbicom(file1, file2):

    try: 
        # Cargar el archivo de Excel usando xlrd
        workbook = xlrd.open_workbook(file1)
        sheet = workbook.sheet_by_index(0)
        workbook2 = xlrd.open_workbook(file2)
        sheet2 = workbook2.sheet_by_index(0)

        # Extraer la información necesaria del reporte
        fecha = sheet.cell_value(11, 11).split()[0]  # Celda L12
        
        km_recorridos = float(sheet.cell_value(20, 12))  # Celda M20
    
        num_excesos = int(sheet.cell_value(20, 21))  # Celda V20
    
        placa = sheet.cell_value(13, 24).split(' - ')[1].replace('(', '').replace(')', '') # Celda Y14, quitando el texto adicional
        

        # Calcular día trabajado y preoperacional
        dia_trabajado = 1 if km_recorridos > 0 else 0
        preoperacional = 1 if dia_trabajado == 1 else None

        # Extraer el número de desplazamientos

    
        num_desplazamientos = 0
        for i in range(17, sheet2.nrows - 1):  # Empezamos en la fila 18 y ajustamos para no contar la última fila
            if sheet2.cell_value(i, 10) != '':  # Columna K es el indice 10
                num_desplazamientos += 1

        # Crear el diccionario con los datos extraídos
        datos_extraidos = {
            'placa': placa,
            'fecha': fecha,
            'km_recorridos': km_recorridos,
            'dia_trabajado': dia_trabajado,
            'preoperacional': preoperacional,
            'num_excesos': num_excesos,
            'num_desplazamientos': num_desplazamientos,
        }

        return [datos_extraidos]
    except Exception as e:
        print('Archivos incorrectos o faltantes UBICOM')
        return []
print("AAAAAAA")
# Extraer los datos de los informes de Wialon.

def extraerWialon(file_path1, file_path2, file_path3):

    try:

        datos_extraidos = []
        
        for file_path in [file_path1, file_path2, file_path3]:
            xl = pd.ExcelFile(file_path)
            
            # Extraer placa y fecha siempre
            if 'Statistics' in xl.sheet_names:
                statistics_df = xl.parse('Statistics', header=None)
                placa = statistics_df.iloc[0, 1]  # Celda B1
                fecha = statistics_df.iloc[1, 1].split()[0].replace('.', '/')  # Celda B2
                fecha_formateada = pd.to_datetime(fecha).strftime('%d/%m/%Y')
            
            
            # Verificar si el archivo tiene datos
            if 'Statistics' in xl.sheet_names and 'Excesos de velocidad' in xl.sheet_names and 'Cronología' in xl.sheet_names:
                km_recorridos = int(statistics_df.iloc[7, 1])  # Celda B8, quitando 'km'
                excesos_df = xl.parse('Excesos de velocidad', header=None)
                num_excesos = len(excesos_df) - 1  # Descontar la fila de encabezado
                crono = xl.parse('Cronología')
                
                # Extraer número de desplazamientos
                desplazamientos = 0
                for x in crono['Tipo'].to_list():
                    if x == 'Trip':
                        desplazamientos += 1
                
                # Calcular día trabajado y preoperacional
                dia_trabajado = 1 if km_recorridos > 0 else 0
                preoperacional = 1 if dia_trabajado == 1 else 0
                
                # Crear el diccionario con los datos extraídos
                datos = {
                    'placa': placa,
                    'fecha': fecha_formateada,
                    'km_recorridos': km_recorridos,
                    'dia_trabajado': dia_trabajado,
                    'preoperacional': preoperacional,
                    'num_excesos': num_excesos,
                    'num_desplazamientos': desplazamientos
                }
            else:
                # Si el archivo no tiene datos, llenar con ceros pero usar placa y fecha
                datos = {
                    'placa': placa,
                    'fecha': fecha_formateada,
                    'km_recorridos': 0,
                    'dia_trabajado': 0,
                    'preoperacional': 0,
                    'num_excesos': 0,
                    'num_desplazamientos': 0
                }
            
            datos_extraidos.append(datos)
        
        return datos_extraidos
    
    except Exception as e: 
        print('Archivos incorrectos o faltantes WIALON')
        return []
print("AAAABBBAAA")
# Infractores diario Ubicar
lugarDescargasUbicar = os.getcwd() + r"\outputUbicar"
print("AABAAAAA")
archivos = glob.glob(os.path.join(lugarDescargasUbicar, '*.xlsx'))
print("AAAAABAA")

def main():
    from forms.ubicarForm import rpaUbicar,archivoUbicar1,archivoUbicar2,archivoUbicar3

    print("AAAABBBBAAA")
    print(archivoUbicar3)
    print(archivoUbicar1)
    print(extraerUbicar(archivoUbicar1,archivoUbicar3))

if __name__ == '__main__':
    main()