import pandas as pd
import openpyxl
import re
import xlrd
from datetime import datetime
from persistence.archivoExcel import OdomIturan, odomUbicar
import sqlalchemy
from sqlalchemy import create_engine, text, Table, update
from sqlalchemy.orm import sessionmaker
from persistence.archivoExcel import infracTodos
from persistence.archivoExcel import xlsx

# ACTUALIZAR SEGUIMIENTO EN LA BASE DE DATOS 'vehiculos'

# Las siguientes funciones extraen la información de cada uno de los archivos de las 6 plataformas y lo reunen todo en un único dataframe para actualizar la tabla 'seguimiento'.

    # Ituran 

def sqlIturan(file1, file2):
    # Cargar el archivo csv
    itu = pd.read_csv(file1)[['NICK_NAME', 'TOTAL_TRIP_DISTANCE', 'TOTAL_NUMBER_OF_TRIPS']]
    itu2 = pd.read_csv(file2)
    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    fecha = datetime.now().strftime('%d/%m/%Y') + ' ' + current_time
    
    # Cambiar el nombre de las columnas
    itu = itu.rename(columns={
        'NICK_NAME': 'placa',
        'TOTAL_TRIP_DISTANCE': 'km_recorridos',
        'TOTAL_NUMBER_OF_TRIPS': 'num_desplazamientos'
    })
    
    # Agregar columna 'fecha'
    itu['fecha'] = fecha
    
    # Agregar columnas 'dia_trabajado' y 'preoperacional'
    itu['dia_trabajado'] = itu['km_recorridos'].apply(lambda x: 1 if x > 0 else 0)
    itu['preoperacional'] = itu['dia_trabajado'].apply(lambda x: 1 if x == 1 else '-')
    
    # Calcular el número de excesos de velocidad y crear DataFrame para fusiones
    excesos = itu2[itu2['TOP_SPEED'] > 80].groupby('V_NICK_NAME').size().reset_index(name='num_excesos')
    excesos = excesos.rename(columns={'V_NICK_NAME': 'placa'})
    
    # Unir el DataFrame de excesos con el DataFrame itu
    itu = itu.merge(excesos, on='placa', how='left')
    
    # Reemplazar los valores NaN en la columna num_excesos por 0
    itu['num_excesos'] = itu['num_excesos'].fillna(0).astype(int)

    itu['proveedor'] = 'Ituran'
    
    # Convertir el DataFrame filtrado a un diccionario sin incluir el índice
    datos_extraidos = itu.to_dict(orient='records')
    
    return datos_extraidos

    # MDVR

def sqlMDVR(file1, file2): #file1 es el informe general, file2 es el informe de paradas (para determinar los desplazamientos)

# Cargar el archivo de Excel usando xlrd
    workbook = xlrd.open_workbook(file1)
    sheet = workbook.sheet_by_index(0)
    file2x = xlsx(file2)
    workbook2 = openpyxl.load_workbook(file2x)
    sheet2 = workbook2.active


    # Extraer la información necesaria del reporte
    placa_completa = sheet.cell_value(0, 1)  # A1 es (0, 1)
    placa = placa_completa.replace('-', '')  # Quitar el guion de la placa
    fecha = sheet.cell_value(1, 1).split()[0]  # A2 es (1, 1)
    km_recorridos = float(sheet.cell_value(3, 1).replace(' Km', ''))  # A5 es (4, 1)
    dia_trabajado = 1 if km_recorridos > 0 else 0
    preoperacional = 1 if dia_trabajado == 1 else '-'
    num_excesos = int(sheet.cell_value(8, 1))  # A9 es (8, 1)

    #Contar número de desplazamientos
    num_desplazamientos = 0
    for i in range(1, sheet2.max_row + 1):
        if sheet2.cell(row=i, column=2).value == 'Movimiento':
            num_desplazamientos += 1
    
    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    # Crear el diccionario con los datos extraídos
    datos_extraidos = {
        'placa': placa,
        'fecha': fecha + ' ' + current_time,
        'km_recorridos': km_recorridos,
        'dia_trabajado': dia_trabajado,
        'preoperacional': preoperacional,
        'num_excesos': num_excesos,
        'num_desplazamientos': num_desplazamientos,
        'proveedor': 'MDVR'
    }

    return [datos_extraidos]

    # Ubicar

def sqlUbicar(file1, file2): # file 1 es el informe general, file2 es el informe de paradas (para determinar los desplazamientos)
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(file1)
    workbook2 = openpyxl.load_workbook(file2)

    sheet = workbook.active

    sheet2 = workbook2.active

    # Extraer la información necesaria del reporte
    placa_completa = sheet.cell(row=1, column=2).value
    # Extraer la parte relevante de la placa
    placa = placa_completa.split()[1] + placa_completa.split()[2]
    fecha = sheet.cell(row=2, column=2).value.split()[0].replace('-','/')
    km_recorridos = float(sheet.cell(row=4, column=2).value.replace(' Km', ''))
    dia_trabajado = 1 if km_recorridos > 0 else 0
    preoperacional = 1 if dia_trabajado == 1 else '-'
    numExcesos = sheet.cell(row=9, column=2).value

    #Contar número de desplazamientos
    num_desplazamientos = 0
    for i in range(5, sheet2.max_row + 1):
         if sheet2.cell(row=i, column=1).value == 'Movimiento':
            num_desplazamientos += 1

    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    # Crear el diccionario con los datos extraídos
    datos_extraidos = {
        'placa': placa,
        'fecha': fecha + '' + current_time,
        'km_recorridos': km_recorridos,
        'dia_trabajado': dia_trabajado,
        'preoperacional': preoperacional,
        'num_excesos': numExcesos,
        'num_desplazamientos': num_desplazamientos,
        'proveedor': 'Ubicar'
    }

    return [datos_extraidos]

    # Ubicom

def sqlUbicom(file1, file2):
    # Cargar el archivo de Excel usando xlrd
    workbook = xlrd.open_workbook(file1)
    sheet = workbook.sheet_by_index(0)
    workbook2 = xlrd.open_workbook(file2)
    sheet2 = workbook2.sheet_by_index(0)

    # Extraer la información necesaria del reporte
    fecha = sheet.cell_value(11, 28).split()[0]  # Celda AC12
    
    km_recorridos = float(sheet.cell_value(20, 12))  # Celda M20
  
    num_excesos = int(sheet.cell_value(20, 21))  # Celda V20
  
    placa = sheet.cell_value(13, 24).split(' - ')[1].replace('(', '').replace(')', '') # Celda Y14, quitando el texto adicional
    

    # Calcular día trabajado y preoperacional
    dia_trabajado = 1 if km_recorridos > 0 else 0
    preoperacional = 1 if dia_trabajado == 1 else '-'

    # Extraer el número de desplazamientos
   
    num_desplazamientos = 0
    for i in range(17, sheet2.nrows - 1):  # Empezamos en la fila 18 y ajustamos para no contar la última fila
        if sheet2.cell_value(i, 10) != '':  # Columna K es el indice 10
            num_desplazamientos += 1

    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    # Crear el diccionario con los datos extraídos
    datos_extraidos = {
        'placa': placa,
        'fecha': fecha + ' ' + current_time,
        'km_recorridos': km_recorridos,
        'dia_trabajado': dia_trabajado,
        'preoperacional': preoperacional,
        'num_excesos': num_excesos,
        'num_desplazamientos': num_desplazamientos,
        'proveedor': 'Ubicom'
    }

    return [datos_extraidos]

    # Securitrac

def sqlSecuritrac(file):
    try:
    # Cargar el archivo de Excel usando pandas
        df = pd.read_excel(file)

        # Diccionario para almacenar los datos por placa
        datos_por_placa = {}

        current_datetime = datetime.now()
        current_time = current_datetime.strftime("%H:%M:%S")
        
        for index, row in df.iterrows():
            placa = row['NROMOVIL']
            evento = row['EVENTO']
            kilometros = float(row['KILOMETROS'])
            fecha = row['FECHAGPS']

        
            fecha_formateada = pd.to_datetime(fecha).strftime('%d/%m/%Y')

            if placa not in datos_por_placa:
                datos_por_placa[placa] = {
                    'placa': placa,
                    'fecha': fecha_formateada + ' ' + current_time, 
                    'km_recorridos': 0,
                    'num_excesos': 0,
                    'num_desplazamientos': 0,
                    'proveedor': 'Securitrac'
                }
            datos_por_placa[placa]['km_recorridos'] += kilometros
            if evento == 'Exc. Velocidad':
                datos_por_placa[placa]['num_excesos'] += 1
            datos_por_placa[placa]['num_desplazamientos'] += 1

        for placa, datos in datos_por_placa.items():
            datos['dia_trabajado'] = 1 if datos['km_recorridos'] > 0 else 0
            datos['preoperacional'] = 1 if datos['dia_trabajado'] == 1 else '-'

        datos = []
        for x in datos_por_placa.keys():
            datos.append(datos_por_placa[x])

        return datos
    except Exception as e:
        print('Archivos incorrectos o faltantes Securitrac')
        return []

    # Wialon

def sqlWialon(file1, file2, file3):
    datos_extraidos = []

    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")
    
    for file in [file1, file2, file3]:
        xl = pd.ExcelFile(file)
        
        # Extraer placa y fecha siempre
        if 'Statistics' in xl.sheet_names:
            statistics_df = xl.parse('Statistics', header=None)
            placa = statistics_df.iloc[0, 1]  # Celda B1
            fecha = statistics_df.iloc[1, 1].split()[0].replace('.', '/')  # Celda B2
            fecha_formateada = pd.to_datetime(fecha).strftime('%d/%m/%Y')
        
        
        # Verificar si el archivo tiene datos
        if 'Statistics' in xl.sheet_names and 'Excesos de velocidad' in xl.sheet_names and 'Cronología' in xl.sheet_names:
            km_recorridos = int(statistics_df.iloc[7, 1])  # Celda B8, quitando 'km'
            excesos_df = xl.parse('Excesos de velocidad', header=None)
            num_excesos = len(excesos_df) - 1  # Descontar la fila de encabezado
            crono = xl.parse('Cronología')
            
            # Extraer número de desplazamientos
            desplazamientos = 0
            for x in crono['Tipo'].to_list():
                if x == 'Trip':
                    desplazamientos += 1
            
            # Calcular día trabajado y preoperacional
            dia_trabajado = 1 if km_recorridos > 0 else 0
            preoperacional = 1 if dia_trabajado == 1 else None
            
            # Crear el diccionario con los datos extraídos
            datos = {
                'placa': placa,
                'fecha': fecha_formateada + ' ' + current_time,
                'km_recorridos': km_recorridos,
                'dia_trabajado': dia_trabajado,
                'preoperacional': preoperacional,
                'num_excesos': num_excesos,
                'num_desplazamientos': desplazamientos,
                'proveedor': 'Wialon'
            }

        else:
            # Si el archivo no tiene datos, llenar con ceros pero usar placa y fecha
            datos = {
                'placa': placa,
                'fecha': fecha_formateada + ' ' + current_time,
                'km_recorridos': 0,
                'dia_trabajado': 0,
                'preoperacional': None,
                'num_excesos': 0,
                'num_desplazamientos': 0,
                'proveedor': 'Wialon'
            }
        
        datos_extraidos.append(datos)
    
    return datos_extraidos

    # Crear DF

def ejecutarTodasExtraccionesSQL(file_ituran, file_ituran2, file_MDVR1, file_MDVR2, file_Ubicar1, file_Ubicar2, file_Ubicom1, file_Ubicom2, file_Securitrac, file_Wialon1, file_Wialon2, file_Wialon3):
    # Ejecutar cada función de extracción con los archivos proporcionados
    datos_mdvr = sqlMDVR(file_MDVR1, file_MDVR2)
    datos_ituran = sqlIturan(file_ituran, file_ituran2)
    datos_securitrac = sqlSecuritrac(file_Securitrac)
    datos_wialon = sqlWialon(file_Wialon1, file_Wialon2, file_Wialon3)
    datos_ubicar = sqlUbicar(file_Ubicar1, file_Ubicar2)
    datos_ubicom = sqlUbicom(file_Ubicom1, file_Ubicom2)

    # Unir todas las listas en una sola lista final
    lista_final = datos_ituran + datos_mdvr + datos_ubicar + datos_ubicom + datos_securitrac + datos_wialon 

    df_final = pd.DataFrame(lista_final)

    df_final.rename(columns={
    'placa': 'placa',
    'fecha': 'fecha',
    'km_recorridos': 'kmRecorridos',
    'dia_trabajado': 'diaTrabajado',
    'preoperacional': 'preoperacional',
    'num_excesos': 'numExcesos',
    'num_desplazamientos': 'numDesplazamientos',
    'proveedor': 'proveedor'
    }, inplace=True)

    ordenColumnas = [
    'placa',
    'kmRecorridos',
    'numDesplazamientos',
    'diaTrabajado',
    'preoperacional',
    'numExcesos',
    'proveedor',
    'fecha'
    ]

    # Reordenar las columnas de df_final
    df_final = df_final[ordenColumnas]

    return df_final

    # Actualizar 'seguimiento'

def actualizarSeguimientoSQL(file_ituran, file_ituran2, file_MDVR1, file_MDVR2, file_Ubicar1, file_Ubicar2, file_Ubicom1, file_Ubicom2, file_Securitrac, file_Wialon1, file_Wialon2, file_Wialon3):

    df_seguimiento = ejecutarTodasExtraccionesSQL(
    file_ituran,
    file_ituran2, 
    file_MDVR1, 
    file_MDVR2, 
    file_Ubicar1, 
    file_Ubicar2,
    file_Ubicom1, 
    file_Ubicom2,
    file_Securitrac,
    file_Wialon1,
    file_Wialon2, 
    file_Wialon3
    )

    user = 'root'
    password = 'Gatitos24'  
    host = '127.0.0.1'
    port = '3306'
    schema = 'vehiculos'

    engine = create_engine(f'mysql+mysqlconnector://{user}:{password}@{host}:{port}/{schema}')

     # Insertar los datos en la tabla seguimiento
    df_seguimiento.to_sql('seguimiento', con=engine, if_exists='append', index=False)

    print("Datos insertados correctamente en la tabla seguimiento.")


# Actualizar infractores (Esto ya está en otra parte, me toca moverlo acá)

def actualizarInfractoresSQL(fileIturan, fileMDVR, fileUbicar, fileWialon, fileWialon2, fileWialon3, fileSecuritrac):
    # Obtener todas las infracciones combinadas
    todos_registros = infracTodos(fileIturan, fileMDVR, fileUbicar, fileWialon, fileWialon2, fileWialon3, fileSecuritrac)
    df_infractores = pd.DataFrame(todos_registros)

    # Convertir la columna 'FECHA' a datetime y luego a string con el formato correcto
    df_infractores['FECHA'] = pd.to_datetime(df_infractores['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%Y-%m-%d')

    # Renombrar las columnas a camelCase
    df_infractores.columns = [
        'placa', 
        'tiempoDeExceso', 
        'duracionEnKmDeExceso', 
        'velocidadMaxima', 
        'proyecto', 
        'rutaDeExceso', 
        'conductor', 
        'fecha'
    ]

    # Conexión a la base de datos MySQL
    engine = sqlalchemy.create_engine('mysql+mysqlconnector://root:Gatitos24@127.0.0.1:3306/vehiculos')

    try:
        # Leer los datos existentes en la tabla 'infractores'
        df_existente = pd.read_sql_table('infractores', con=engine)

        # Concatenar los datos existentes con los nuevos datos
        df_final = pd.concat([df_existente, df_infractores], ignore_index=True)

        # Eliminar duplicados opcionalmente
        df_final = df_final.drop_duplicates(subset=['placa', 'fecha', 'tiempoDeExceso', 'duracionEnKmDeExceso', 'velocidadMaxima', 'proyecto', 'rutaDeExceso', 'conductor'])

    except ValueError:  # Si la tabla no existe aún, usa los nuevos datos directamente
        df_final = df_infractores

    # Exportar el DataFrame combinado a la tabla 'infractores'
    df_final.to_sql(name='infractores', con=engine, if_exists='replace', index=False)

    print("Datos actualizados en la tabla 'infractores'")
    return df_final

# Actualizar Kilometraje. 

def actualizarKilometraje(file_ituran, file_ubicar):
    todos_registros = OdomIturan(file_ituran) + odomUbicar(file_ubicar)
    df_odometro = pd.DataFrame(todos_registros)


    user = 'root'
    password = 'Gatitos24'
    host = '127.0.0.1'
    port = '3306'
    schema = 'vehiculos'
    tabla = 'carro'

    # Crear la cadena de conexión usando las variables
    engine = create_engine(f'mysql+mysqlconnector://{user}:{password}@{host}:{port}/{schema}')

    Session = sessionmaker(bind=engine)
    session = Session()

    consulta_sql = text(f'''
    UPDATE {tabla}
    SET {'Kilometraje'} = :nuevo_valor
    WHERE {'Placas'} = :primary_key_value
    ''')
    for index, placa in df_odometro.iterrows():

        kilometraje = placa['KILOMETRAJE']
        placa1 = placa['PLACA']

        session.execute(consulta_sql, {'nuevo_valor': kilometraje, 'primary_key_value': placa1})
        session.commit()
    session.close() 
    