import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from persistence.archivoExcel import FuncionalidadExcel


class Extracciones:
    def __init__(self):
        pass

        
    def crear_excel(self, archivoMDVR1, archivoMDVR2, archivoIturan1, archivoIturan2, archivoSecuritrac, archivoWialon1, archivoWialon2, archivoWialon3, archivoUbicar1, archivoUbicar2, archivoUbicom1, archivoUbicom2, output_file):
        # Ejecutar todas las extracciones


        # Ejecutar cada función de extracción con los archivos proporcionados
        datosMDVR = FuncionalidadExcel().extraerMDVR(archivoMDVR1, archivoMDVR2)
        datosIturan = FuncionalidadExcel().extraerIturan(archivoIturan1 , archivoIturan2)
        datosSecuritrac = FuncionalidadExcel().extraerSecuritrac(archivoSecuritrac)
        datosWialon = FuncionalidadExcel().extraerWialon(archivoWialon1, archivoWialon2, archivoWialon3)
        datosUbicar = FuncionalidadExcel().extraerUbicar(archivoUbicar1, archivoUbicar2)
        datosUbicom = FuncionalidadExcel().extraerUbicom(archivoUbicom1, archivoUbicom2)

        # Unir todas las listas en una sola lista final
        nuevos_datos = datosMDVR + datosIturan + datosSecuritrac + datosWialon + datosUbicar + datosUbicom


        # Convertir la lista de nuevos datos a DataFrame
        df_nuevos = pd.DataFrame(nuevos_datos)

        if not os.path.exists(output_file):
            # Si el archivo no existe, crear el DataFrame inicial con el formato deseado
            placas = df_nuevos['placa'].unique()
            fechas = pd.date_range(start='2024-01-01', periods=365, freq='D')  # Ajustar el rango de fechas según sea necesario

            rows = []
            for placa in placas:
                rows.append({'PLACA': placa, 'SEGUIMIENTO': 'Nº Excesos'})
                rows.append({'PLACA': placa, 'SEGUIMIENTO': 'Nº Desplazamiento'})
                rows.append({'PLACA': placa, 'SEGUIMIENTO': 'Día Trabajado'})
                rows.append({'PLACA': placa, 'SEGUIMIENTO': 'Preoperacional'})
                rows.append({'PLACA': placa, 'SEGUIMIENTO': 'Km recorridos'})

            df_formato = pd.DataFrame(rows)

            for fecha in fechas:
                mes_dia = fecha.strftime('%d/%m')
                df_formato[mes_dia] = ''

            # Guardar el DataFrame en un archivo Excel
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                df_formato.to_excel(writer, sheet_name='Seguimiento', index=False)

        else:
            # Leer el archivo existente
            book = load_workbook(output_file)
            if 'Seguimiento' in book.sheetnames:
                sheet = book['Seguimiento']
                self.df_existente = pd.read_excel(output_file, sheet_name='Seguimiento')
            else:
                sheet = book.create_sheet('Seguimiento')
                self.df_existente = pd.DataFrame()

            # Rellenar los datos en el DataFrame con el formato deseado
            for _, row in df_nuevos.iterrows():
                fecha = pd.to_datetime(row['fecha'], format='%d/%m/%Y')
                dia = fecha.strftime('%d/%m')
                placa = row['placa']

                self.df_existente.loc[(self.df_existente['PLACA'] == placa) & (self.df_existente['SEGUIMIENTO'] == 'Nº Excesos'), dia] = row['num_excesos']
                self.df_existente.loc[(self.df_existente['PLACA'] == placa) & (self.df_existente['SEGUIMIENTO'] == 'Nº Desplazamiento'), dia] = row['num_desplazamientos']
                self.df_existente.loc[(self.df_existente['PLACA'] == placa) & (self.df_existente['SEGUIMIENTO'] == 'Día Trabajado'), dia] = row['dia_trabajado']
                self.df_existente.loc[(self.df_existente['PLACA'] == placa) & (self.df_existente['SEGUIMIENTO'] == 'Preoperacional'), dia] = row['preoperacional']
                self.df_existente.loc[(self.df_existente['PLACA'] == placa) & (self.df_existente['SEGUIMIENTO'] == 'Km recorridos'), dia] = row['km_recorridos']


            # Rellenar con 0's espacios en blanco. Esto puede ser necesario cambiarlo dependiendo de cómo el read_excel interprete los valores vacios del excel (NaN o ''). Lo voy a dejar comentado.
            current_date = pd.to_datetime('today').strftime('%d/%m')
            for col in self.df_existente.columns[2:]:  # Saltar 'PLACA' y 'SEGUIMIENTO'.
                if pd.to_datetime(col, format='%d/%m') < pd.to_datetime(current_date, format='%d/%m'):
                    self.df_existente[col].replace('', 0, inplace=True)


            # Escribir los datos actualizados en la hoja 'seguimiento'
            for r_idx, row in enumerate(dataframe_to_rows(self.df_existente, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Guardar el archivo Excel
            book.save(output_file)
            return self.df_existente

    
    def actualizarInfractores(self, file_seguimiento, file_Ituran, file_MDVR, file_Ubicar, fileWialon1, fileWialon2, fileWialon3, file_Securitrac):
    
        # Obtener todas las infracciones combinadas
        registros_ituran = FuncionalidadExcel().infracIturan(file_Ituran)
        registros_mdvr = FuncionalidadExcel().infracMDVR(file_MDVR)
        registros_ubicar = FuncionalidadExcel().infracUbicar(file_Ubicar)
        registros_wialon = FuncionalidadExcel().infracWialon(fileWialon1)
        registros_wialon2 = FuncionalidadExcel().infracWialon(fileWialon2)
        registros_wialon3 = FuncionalidadExcel().infracWialon(fileWialon3)
        registros_securitrac = FuncionalidadExcel().infracSecuritrac(file_Securitrac)

        # Combinar todos los resultados en una sola lista
        todos_registros = (
            registros_ituran + registros_mdvr + registros_ubicar + registros_wialon + registros_wialon2 + registros_wialon3 +registros_securitrac
        )

        df_infractores = pd.DataFrame(todos_registros)

        # Convertir la columna 'FECHA' a datetime y luego a string con el formato correcto
        df_infractores['FECHA'] = pd.to_datetime(df_infractores['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')

        df_infractores = df_infractores[(df_infractores['VELOCIDAD MÁXIMA'] > 80) & (df_infractores['TIEMPO DE EXCESO'] > 20)]

        try:
            # Cargar el archivo existente y añadir una nueva hoja
            with pd.ExcelWriter(file_seguimiento, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Verificar si la hoja 'Infractores' ya existe
                if 'Infractores' in writer.book.sheetnames:
                    # Leer la hoja existente en un DataFrame
                    df_existente = pd.read_excel(file_seguimiento, sheet_name='Infractores')
                    # Concatenar los datos existentes con los nuevos datos
                    df_final = pd.concat([df_existente, df_infractores], ignore_index=True)
                else:
                    # Si la hoja no existe, simplemente usar los nuevos datos
                    df_final = df_infractores

                # Escribir el DataFrame en la hoja 'Infractores'
                df_final.to_excel(writer, sheet_name='Infractores', index=False)

            print(f"Agregado como hoja 'Infractores' en {file_seguimiento}")

        except Exception as e:
            print(f"Error al actualizar el archivo Excel: {e}")


    def actualizarOdom(self, file_seguimiento, file_ituran, file_ubicar):
        # Obtener todos los odómetros combinados
        todos_registros = FuncionalidadExcel().OdomIturan(file_ituran) + FuncionalidadExcel().odomUbicar(file_ubicar)
        df_odometros = pd.DataFrame(todos_registros)

        # Cargar el archivo existente y añadir una nueva hoja, sobreescribiendo si ya existe
        with pd.ExcelWriter(file_seguimiento, engine='openpyxl', mode='a') as writer:
            # Verificar si la hoja ya existe
            if 'Odómetro' in writer.book.sheetnames:
                # Eliminar la hoja existente
                std = writer.book['Odómetro']
                writer.book.remove(std)
            # Escribir el DataFrame en una nueva hoja llamada 'Odometro'
            df_odometros.to_excel(writer, sheet_name='Odómetro', index=False)


    def actualizarIndicadoresTotales(self, df_diario, file_seguimiento):
        # Convertir la columna 'FECHA' a datetime y luego formatear para quitar la hora
        df_diario['FECHA'] = pd.to_datetime(df_diario['FECHA'], format='%Y-%m-%d')

        # Escribir el DataFrame en la hoja 'Indicadores Totales', reemplazando si existe
        with pd.ExcelWriter(file_seguimiento, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_diario.to_excel(writer, sheet_name='Indicadores Totales', index=False)

        print(f"Agregado como hoja 'Indicadores Totales' en {file_seguimiento}")

    
    def actualizarIndicadores(self, df_diario, df_exist, file_seguimiento):


        # Crear df_diario y df_hist
    

        # Calcular los cuatro indicadores
        df_EJL = FuncionalidadExcel().calcular_EJL(df_diario)
        df_GVE = FuncionalidadExcel().calcular_GVE(df_diario, df_exist)
        df_ELVL = FuncionalidadExcel().calcular_ELVL(df_diario)
        df_IDP = FuncionalidadExcel().calcular_IDP(df_diario)

        # Crear una lista de DataFrames
        dfs = [df_EJL, df_GVE, df_ELVL, df_IDP]

        # Crear el archivo si no existe
        if not os.path.exists(file_seguimiento):
            with pd.ExcelWriter(file_seguimiento, engine='xlsxwriter') as writer:
                # Crea un archivo de Excel vacío
                pd.DataFrame().to_excel(writer)

        # Cargar el archivo existente
        book = load_workbook(file_seguimiento)

        # Crear o seleccionar la hoja de trabajo 'Indicadores'
        if 'Indicadores' in book.sheetnames:
            del book['Indicadores']
        sheet = book.create_sheet('Indicadores')

        # Escribir cada DataFrame en la ubicación especificada con un espacio de 1 fila entre ellos
        start_row = 1
        for df in dfs:
            df = df.reset_index()
            df.rename(columns={'index': 'Indicador'}, inplace=True)
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(row):
                    cell = sheet.cell(row=r_idx, column=c_idx + 1, value=value)
                    if r_idx == start_row or c_idx == 0:  # Formato a los encabezados de los meses y la columna 'Indicador'
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            start_row += len(df) + 2  # Incrementar la fila inicial para el siguiente DataFrame (1 fila de espacio + 1 fila de encabezado)

        # Guardar el archivo
        book.save(file_seguimiento)


    def dfDiario(self, df_exist):
        sumas_diario = {
            'FECHA': [], 
            'KILOMETROS RECORRIDOS': [], 
            'EXCESOS VELOCIDAD': [], 
            'DESPLAZAMIENTOS': [], 
            'DÍA TRABAJADO': [], 
            'PREOPERACIONAL': []
        }

        date_columns = df_exist.columns[2:]

        
        for date in date_columns:
            sumas_diario['FECHA'].append(date)
            sumas_diario['KILOMETROS RECORRIDOS'].append(df_exist[df_exist['SEGUIMIENTO'] == 'Km recorridos'][date].sum())
            sumas_diario['EXCESOS VELOCIDAD'].append(df_exist[df_exist['SEGUIMIENTO'] == 'Nº Excesos'][date].sum())
            sumas_diario['DESPLAZAMIENTOS'].append(df_exist[df_exist['SEGUIMIENTO'] == 'Nº Desplazamiento'][date].sum())
            sumas_diario['DÍA TRABAJADO'].append(df_exist[df_exist['SEGUIMIENTO'] == 'Día Trabajado'][date].sum())
            sumas_diario['PREOPERACIONAL'].append(df_exist[df_exist['SEGUIMIENTO'] == 'Preoperacional'][date].sum())

        self.df_diario = pd.DataFrame(sumas_diario)
        self.df_diario['FECHA'] = pd.to_datetime(self.df_diario['FECHA'] + '/2024', format='%d/%m/%Y').dt.strftime('%Y-%m-%d')  # Toca ajustar el año según lo necesitemos.
        self.df_diario['FECHA'] = pd.to_datetime(self.df_diario['FECHA'])

        return self.df_diario


    def fueraLaboralTodos(self, rutasLaboral):
    
        all_results = []
        all_results.extend(self.fueraLaboralSecuritrac(rutasLaboral['securitrac']))
        all_results.extend(self.fueraLaboralMDVR(rutasLaboral['mdvr']))
        all_results.extend(self.fueraLaboralUbicar(rutasLaboral['ubicar']))
        all_results.extend(self.fueraLaboralIturan(rutasLaboral['ituran']))
    
        for file_path in rutasLaboral['wialon']:
            all_results.extend(self.fueraLaboralWialon(file_path))
        
        self.todosDF = pd.DataFrame(all_results)
        self.todosDF['fecha'] = pd.to_datetime(self.todosDF['fecha'], format='%d/%m/%Y %H:%M').dt.strftime('%Y-%m-%d %H:%M:%S')

        return self.todosDF    