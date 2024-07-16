import pandas as pd
import openpyxl
import re
import xlrd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from tkinter import messagebox

from persistence.indicadores import Indicadores
from persistence.extraerExcel import ExtraerExcel
from db.conexionDB import conexionDB

class Historico(conexionDB):

    def __init__(self):
        super().__init__()
        self.extraer = ExtraerExcel()
        self.indicador = Indicadores()

    # Esto solo se debería correr una ÚNICA vez con los datos históricos, y solo después de haber creado el excel. 
    # Actualizar históricos
    def agregar_datos_historicos(self, file_seguimiento, file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac, file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
        # Leer el archivo de seguimiento existente
        df_seguimiento = pd.read_excel(file_seguimiento)
        
        # Ejecutar las funciones de historial y recopilar los datos
        datos_ubicar = self.extraer.histUbicarExcel(file_ubicar1, file_ubicar2)
        datos_MDVR = self.extraer.histMDVRExcel(file_mdvr1, file_mdvr2)
        datos_ubicom = self.extraer.histUbicomExcel(file_ubicom1, file_ubicom2)
        datos_securitrac = self.extraer.histSecuritracExcel(file_securitrac)
        datos_wialon1 = self.extraer.histWialonExcel(file_wialon1)
        datos_wialon2 = self.extraer.histWialonExcel(file_wialon2)
        datos_wialon3 = self.extraer.histWialonExcel(file_wialon3)
        datos_ituran = self.extraer.histIturanExcel(file_ituran1, file_ituran2)
        
        # Concatenar todos los datos históricos
        data_historica = datos_ubicar + datos_MDVR + datos_ubicom + datos_securitrac + datos_wialon1 + datos_wialon2 + datos_wialon3 + datos_ituran
        
        # Convertir los datos extraídos a un DataFrame
        df_hist = pd.DataFrame(data_historica)
        
        # Corregir las fechas NaN copiando el valor de la columna 'Fecha'
        df_hist['fecha'] = df_hist.apply(lambda row: row['Fecha'] if pd.isna(row['fecha']) else row['fecha'], axis=1)

        # Asegurar que la fecha esté en el formato correcto y tenga solo día y mes
        df_hist['fecha'] = pd.to_datetime(df_hist['fecha'], format='%d/%m/%Y').dt.strftime('%d/%m')
        
        print(f"Contenido de df_hist:")
        print(df_hist.head())
        
        # Rellenar los datos en el DataFrame existente con el formato deseado
        for index, row in df_hist.iterrows():
            fecha = row['fecha']
            placa = row['placa']
            
            if fecha in df_seguimiento.columns:
                if placa in df_seguimiento['PLACA'].values:
                    df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Nº Excesos'), fecha] = row['num_excesos']
                    df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Nº Desplazamiento'), fecha] = row['num_desplazamientos']
                    df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Día Trabajado'), fecha] = row['dia_trabajado']
                    df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Preoperacional'), fecha] = row['preoperacional']
                    df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Km recorridos'), fecha] = row['km_recorridos']
        
        # Guardar el archivo de seguimiento actualizado
        with pd.ExcelWriter(file_seguimiento, engine='openpyxl') as writer:
            df_seguimiento.to_excel(writer, index=False, sheet_name='Seguimiento')
            worksheet = writer.sheets['Seguimiento']
            worksheet.freeze_panes = worksheet['C2']  # Congela la primera fila y las dos primeras columnas

            worksheet.freeze_panes = worksheet['C2']  # Congela la primera fila y las dos primeras columnas

    # Esta es la función que crea por PRIMERA vez la hoja de infractores.
    def crearInfractores(self, file_seguimiento, file_Ituran, file_MDVR, file_Ubicar, file_wialon1, file_wialon2, file_wialon3, file_Securitrac):
        # Obtener todas las infracciones combinadas
        todos_registros = self.extraer.infracTodos(file_Ituran, file_MDVR, file_Ubicar, file_wialon1, file_wialon2, file_wialon3, file_Securitrac)
        df_infractores = pd.DataFrame(todos_registros)

        # Convertir la columna 'FECHA' a datetime y luego a string con el formato correcto
        df_infractores['FECHA'] = pd.to_datetime(df_infractores['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')

        # Cargar el archivo existente y añadir una nueva hoja
        with pd.ExcelWriter(file_seguimiento, engine='openpyxl', mode='a') as writer:
            df_infractores.to_excel(writer, sheet_name='Infractores', index=False)

    def juntarDatosSQL(self, file_ituran1, file_ituran2, file_mdvr1, file_mdvr2, file_ubicar1, file_ubicar2, file_ubicom1, file_ubicom2, file_securitrac, file_wialon1, file_wialon2, file_wialon3):
        # Ejecutar las funciones y recopilar los datos
        datos_ubicar = self.extraer.ubicarHistoricoSQL(file_ubicar1, file_ubicar2)
        datos_MDVR = self.extraer.MDVRHistoricoSQL(file_mdvr1, file_mdvr2)
        datos_ubicom = self.extraer.ubicomHistoricoSQL(file_ubicom1, file_ubicom2)
        datos_securitrac = self.extraer.securitracHistoricoSQL(file_securitrac)
        datos_wialon1 = self.extraer.wialonHistoricoSQL(file_wialon1)
        datos_wialon2 = self.extraer.wialonHistoricoSQL(file_wialon2)
        datos_wialon3 = self.extraer.wialonHistoricoSQL(file_wialon3)
        datos_ituran = self.extraer.ituranHistoricoSQL(file_ituran1, file_ituran2)
        
        # Concatenar todos los datos 
        dataHistorica = datos_ituran + datos_MDVR + datos_ubicar +  datos_ubicom + datos_securitrac + datos_wialon1 + datos_wialon2 + datos_wialon3
        # Convertir los datos extraídos a un DataFrame
        df_hist = pd.DataFrame(dataHistorica)
        df_hist['Fecha'] = df_hist['Fecha'].fillna(df_hist['fecha'])

        # Eliminar la columna 'fecha'
        df_hist.drop(columns=['fecha'], inplace=True)

        df_hist.rename(columns={
        'placa': 'placa',
        'km_recorridos': 'kmRecorridos',
        'num_excesos': 'numExcesos',
        'num_desplazamientos': 'numDesplazamientos',
        'dia_trabajado': 'diaTrabajado',
        'preoperacional': 'preoperacional',
        'proveedor': 'proveedor',
        'Fecha': 'fecha'  
        }, inplace=True)

        df_hist.fillna(0, inplace=True)

        return df_hist

    #### INDICADORES ###
    def crearDfHist(self, file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                    file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
        
        data_historica = []
        
        data_historica.extend(self.extraer.histUbicomExcel(file_ubicom1, file_ubicom2))
        data_historica.extend(self.extraer.histUbicarExcel(file_ubicar1, file_ubicar2))
        data_historica.extend(self.extraer.histMDVRExcel(file_mdvr1, file_mdvr2))
        data_historica.extend(self.extraer.histSecuritracExcel(file_securitrac))
        data_historica.extend(self.extraer.histWialonExcel(file_wialon1))
        data_historica.extend(self.extraer.histWialonExcel(file_wialon2))
        data_historica.extend(self.extraer.histWialonExcel(file_wialon3))
        data_historica.extend(self.extraer.histIturanExcel(file_ituran1, file_ituran2))
        
        df_hist = pd.DataFrame(data_historica)

    # Corregir las fechas NaN copiando el valor de la columna 'Fecha'

        df_hist['fecha'] = df_hist.apply(lambda row: row['Fecha'] if pd.isna(row['fecha']) else row['fecha'], axis=1)

        df_hist['fecha'] = pd.to_datetime(df_hist['fecha'], format='%d/%m/%Y')

        return df_hist
        
    def crear_df_diario(self, file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                    file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
        
        df_hist =  self.crearDfHist(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                    file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2)
        # Asegurarse de que las columnas existan en df_hist y en el formato esperado
        expected_columns = ['placa', 'fecha', 'num_excesos', 'num_desplazamientos', 'dia_trabajado', 'preoperacional', 'km_recorridos']

        for col in expected_columns:
            if col not in df_hist.columns:
                df_hist[col] = None

        df_diario = df_hist.groupby('fecha').agg({
            'km_recorridos': 'sum',
            'num_excesos': 'sum',
            'num_desplazamientos': 'sum',
            'dia_trabajado': 'sum',
            'preoperacional': 'sum'
        }).reset_index()

        # Renombramos las columnas para que coincidan con el formato deseado
        df_diario.columns = ['FECHA', 'KILOMETROS RECORRIDOS', 'EXCESOS VELOCIDAD', 'DESPLAZAMIENTOS', 'DÍA TRABAJADO', 'PREOPERACIONAL']

        # Ordenar el DataFrame por la columna 'FECHA'
        df_diario = df_diario.sort_values(by='FECHA', key=lambda x: pd.to_datetime(x, format='%d/%m/%Y'))

        # Reiniciar el índice si es necesario
        df_diario = df_diario.reset_index(drop=True)

        return df_diario

    def exportar_df_diario(self, file_seguimiento, file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                    file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
        
        df_diario = self.crear_df_diario(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                    file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2)
        # Convertir la columna 'FECHA' a datetime y luego formatear para quitar la hora
        df_diario['FECHA'] = pd.to_datetime(df_diario['FECHA'], format="%d/%m/%Y").dt.strftime('%d/%m/%Y')

        with pd.ExcelWriter(file_seguimiento, mode='a', engine='openpyxl') as writer:
            df_diario.to_excel(writer, sheet_name='Indicadores Totales', index=False)

        print(f"Agregado como hoja 'Indicadores Totales' en {file_seguimiento}")

    def crearIndicadores(self, file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2, file_seguimiento):
        # Crear df_dirio y df_hist
        df_hist = self.crearDfHist(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac, file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2)
        df_diario = self.crear_df_diario(df_hist)
        # Calcular los cuatro indicadores

        df_EJL = self.indicador.calcular_EJL(df_diario)
        df_GVE = self.indicador.calcular_GVE(df_diario, df_hist)
        df_ELVL = self.indicador.calcular_ELVL(df_diario)
        df_IDP = self.indicador.calcular_IDP(df_diario)

        # Crear una lista de DataFrames
        dfs = [df_EJL, df_GVE, df_ELVL, df_IDP]
        
        # Crear el archivo si no existe
        if not os.path.exists(file_seguimiento):
            with pd.ExcelWriter(file_seguimiento, engine='xlsxwriter') as writer:
                # Crea un archivo de Excel vacío
                pd.DataFrame().to_excel(writer)
        
        # Cargar el archivo existente
        book = load_workbook(file_seguimiento)
        
        # Crear o seleccionar la hoja de trabajo 'Indicadores'
        if 'Indicadores' in book.sheetnames:
            sheet = book['Indicadores']
        else:
            sheet = book.create_sheet('Indicadores')
        
        # Escribir cada DataFrame en la ubicación especificada con un espacio de 1 fila entre ellos
        start_row = 1
        for df in dfs:
            df = df.reset_index()
            df.rename(columns={'index': 'Indicador'}, inplace=True)
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row or c_idx == 1:  # Aplicar formato a los encabezados de los meses y la columna 'Indicador'
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            start_row += len(df) + 2  # Incrementar la fila inicial para el siguiente DataFrame (1 fila de espacio + 1 fila de encabezado)
        
        # Guardar el archivo
        book.save(file_seguimiento)

#### Esto probablemente sea mejor ejectutarlo en otro lado y dejar el script solo con las funciones, pero como esto es un script que en teoría solo se va a usar esta vez vale la pena revisar ###


## ESTAS RUTAS TOCA ACTUALIZARLAS CON LAS MÁS RECIENTES

# seguimiento = r"C:\Users\SGI SAS\Desktop\SGI\seguimiento.xlsx"
# mdvr_file1 = r"C:\Users\SGI SAS\Downloads\travel_sheet_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720735191.xls"
# mdvr_file2 = r"C:\Users\SGI SAS\Downloads\overspeeds_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720734995.xlsx"
# ituran_file1 = r"C:\Users\SGI SAS\Downloads\Export.xls"
# ituran_file2 = r"C:\Users\SGI SAS\Downloads\Export(22).xls"
# securitrac_file = r"C:\Users\SGI SAS\Downloads\exported-excel(5).xls"
# wialon_file1 = r"C:\Users\SGI SAS\Downloads\LPN816_INFORME_GENERAL_TM_V1.0_2024-07-11_17-08-35.xlsx"
# wialon_file2 = r"C:\Users\SGI SAS\Downloads\LPN821_INFORME_GENERAL_TM_V1.0_2024-07-11_17-10-13.xlsx"
# wialon_file3 = r"C:\Users\SGI SAS\Downloads\JTV645_INFORME_GENERAL_TM_V1.0_2024-07-11_17-08-06.xlsx"
# ubicar_file1 = r"C:\Users\SGI SAS\Downloads\travel_sheet_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720735468.xls"
# ubicar_file2 = r"C:\Users\SGI SAS\Downloads\overspeeds_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720735429.xlsx"
# ubicom_file1 = r"C:\Users\SGI SAS\Downloads\ReporteDiario.xls"
# ubicom_file2 = r"C:\Users\SGI SAS\Downloads\Estacionados.xls"


# agregar_datos_historicos(seguimiento, ubicar_file1, ubicar_file2, mdvr_file1, mdvr_file2, ubicom_file1, ubicom_file2, securitrac_file, wialon_file1, wialon_file2, wialon_file3, ituran_file1, ituran_file2)