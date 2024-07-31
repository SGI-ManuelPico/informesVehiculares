import pandas as pd
import openpyxl
import re
import xlrd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from datetime import datetime
from sqlalchemy import create_engine
from openpyxl.styles import Font, PatternFill

# Esto solo se debería correr una ÚNICA vez con los datos históricos, y solo después de haber creado el excel. 

# Esto toca ver si lo separamos mejor en dos scripts para que haya mayor claridad.


# Ituran

def histIturan(file1, file2):
    # Leer los archivos Excel
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Convertir las fechas a un formato de fecha datetime
    df1['DT_DRIVE_DATE'] = pd.to_datetime(df1['DT_DRIVE_DATE']).dt.date
    df2['EVENT_START_DAY_TIME'] = pd.to_datetime(df2['EVENT_START_DAY_TIME']).dt.date

    # Crear un DataFrame para almacenar los resultados
    results = []

    # Obtener las placas únicas
    placas = df1['V_NICK_NAME'].unique()

    for placa in placas:
        # Filtrar los datos por placa
        df1_placa = df1[df1['V_NICK_NAME'] == placa]
        df2_placa = df2[df2['V_NICK_NAME'] == placa]

        # Agrupar por fecha para contar los desplazamientos y sumar las distancias
        daily_stats_df1 = df1_placa.groupby('DT_DRIVE_DATE').agg({
            'TRIP_DISTANCE': 'sum',
            'DT_DRIVE_DATE': 'count'  # Esta cuenta el número de desplazamientos
        }).rename(columns={'DT_DRIVE_DATE': 'num_desplazamientos', 'TRIP_DISTANCE': 'km_recorridos'}).reset_index()

        # Agrupar por fecha para contar los excesos
        daily_stats_df2 = df2_placa.groupby('EVENT_START_DAY_TIME').size().reset_index(name='num_excesos')

        # Unir ambos DataFrames por fecha
        daily_stats = pd.merge(daily_stats_df1, daily_stats_df2, left_on='DT_DRIVE_DATE', right_on='EVENT_START_DAY_TIME', how='left').fillna(0)

        for _, row in daily_stats.iterrows():
            fecha_formateada = row['DT_DRIVE_DATE'].strftime('%d/%m/%Y')
            datos_extraidos = {
                'placa': placa,
                'fecha': fecha_formateada,
                'km_recorridos': round(row['km_recorridos'], 2),
                'num_desplazamientos': row['num_desplazamientos'],
                'dia_trabajado': 1 if row['km_recorridos'] > 0 else 0,
                'preoperacional': 1 if row['km_recorridos'] > 0 else None,
                'num_excesos': int(row['num_excesos'])
            }
            results.append(datos_extraidos)

    return results

# MDVR

def histMDVR(file1, file2):
    # Leer el archivo de desplazamientos
    df1 = pd.read_excel(file1, engine='xlrd', header=2, skipfooter=2)
    df1['Fecha'] = pd.to_datetime(df1['Fecha'], format='%d/%m/%Y %H:%M:%S').dt.date
    
    # Limpiar y convertir 'Longitud de ruta' a float
    df1['Longitud de ruta'] = df1['Longitud de ruta'].str.replace(' Km', '').astype(float)
    
    # Contar desplazamientos y sumar kilómetros por día
    desplazamientos_km = df1.groupby('Fecha').agg(
        num_desplazamientos=('Fecha', 'size'),
        km_recorridos=('Longitud de ruta', 'sum')
    ).reset_index()
    
    
    file_path = file2
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Extraer los datos de la hoja y convertirlos en un DataFrame
    data = sheet.values
    next(data) 
    next(data) 

    next(data) 
    

    cols = ["Comienzo", "Fin", "Duración exceso de velocidad", "Velocidad máxima", "Velocidad media", "Posición"]
    data = list(data)
    df2 = pd.DataFrame(data, columns=cols)
    
    # Quitar la última fila
    df2 = df2.iloc[:-1]
    
    # Convertir la columna 'Comienzo' a datetime
    df2['Comienzo'] = pd.to_datetime(df2['Comienzo'], format='%d/%m/%Y %H:%M:%S').dt.date
    
    # Contar excesos de velocidad por día
    excesos_velocidad = df2.groupby('Comienzo').size().reset_index(name='num_excesos')
    
    # Unir los datos de desplazamientos y excesos de velocidad
    datos_historicos = pd.merge(desplazamientos_km, excesos_velocidad, left_on='Fecha', right_on='Comienzo', how='left').fillna(0)
    
    # Añadir la columna fija 'placa'
    datos_historicos['placa'] = 'KSZ298'
    
    # Añadir las columnas faltantes
    datos_historicos['dia_trabajado'] = (datos_historicos['km_recorridos'] > 0).astype(int)
    datos_historicos['preoperacional'] = datos_historicos['dia_trabajado'].replace(0, None)
    
    # Seleccionar y renombrar columnas
    datos_historicos = datos_historicos[['placa', 'Fecha', 'km_recorridos', 'num_desplazamientos', 'dia_trabajado', 'preoperacional', 'num_excesos']]
    datos_historicos = datos_historicos.rename(columns={'Fecha': 'fecha'})
    
    # Convertir la columna 'fecha' a datetime
    datos_historicos['fecha'] = pd.to_datetime(datos_historicos['fecha'])
    
    # Convertir la fecha a formato dd/mm/yyyy
    datos_historicos['fecha'] = datos_historicos['fecha'].dt.strftime('%d/%m/%Y')
    
    return datos_historicos.to_dict(orient='records')

# Ubicar

def histUbicar(file1, file2):
    # Leer el archivo de desplazamientos
    df1 = pd.read_excel(file1, engine='xlrd', header=2, skipfooter=2)
    df1['Fecha'] = df1['Fecha'].apply(lambda x: x.replace('-', '/'))
    df1['Fecha'] = pd.to_datetime(df1['Fecha'], format="%d/%m/%Y %H:%M:%S").dt.date

    # Limpiar y convertir 'Longitud de ruta' a float
    df1['Longitud de ruta'] = df1['Longitud de ruta'].str.replace(' Km', '').astype(float)

    # Contar desplazamientos y sumar kilómetros por día
    desplazamientos_km = df1.groupby('Fecha').agg(
        num_desplazamientos=('Fecha', 'size'),
        km_recorridos=('Longitud de ruta', 'sum')
    ).reset_index()

    # Leer el archivo de excesos de velocidad
    workbook = openpyxl.load_workbook(file2)
    sheet = workbook.active

    # Extraer los datos de la hoja y convertirlos en un DataFrame
    data = sheet.values
    next(data) # Saltar la primera fila de datos (encabezados principales)
    next(data) # Saltar la segunda fila de datos (encabezados secundarios)
    next(data)

    cols = ["Comienzo", "Fin", "Duración exceso de velocidad", "Velocidad máxima", "Velocidad media", "Posición"]
    data = list(data)
    df2 = pd.DataFrame(data, columns=cols)

    # Quitar la última fila
    df2 = df2.iloc[:-1]

    # Convertir la columna 'Comienzo' a datetime
    df2['Comienzo'] = pd.to_datetime(df2['Comienzo'].str.replace('.', '/'), format="%d/%m/%Y %H:%M:%S", errors='coerce').dt.date

    # Contar excesos de velocidad por día
    excesos_velocidad = df2.groupby('Comienzo').size().reset_index(name='num_excesos')

    # Unir los datos de desplazamientos y excesos de velocidad
    datos_historicos = pd.merge(desplazamientos_km, excesos_velocidad, left_on='Fecha', right_on='Comienzo', how='left').fillna(0)

    # Añadir la columna fija 'placa'
    datos_historicos['placa'] = 'JYT682'

    # Añadir las columnas faltantes
    datos_historicos['dia_trabajado'] = (datos_historicos['km_recorridos'] > 0).astype(int)
    datos_historicos['preoperacional'] = (datos_historicos['km_recorridos'] > 0).astype(int)
    datos_historicos.loc[datos_historicos['dia_trabajado'] == 0, 'preoperacional'] = None

    # Seleccionar y renombrar columnas
    datos_historicos = datos_historicos[['placa', 'Fecha', 'km_recorridos', 'num_desplazamientos', 'dia_trabajado', 'preoperacional', 'num_excesos']]
    datos_historicos = datos_historicos.rename(columns={'Fecha': 'fecha'})

    # Convertir la columna 'fecha' a datetime
    datos_historicos['fecha'] = pd.to_datetime(datos_historicos['fecha'])

    # Convertir la fecha a formato dd/mm/yyyy
    datos_historicos['fecha'] = datos_historicos['fecha'].dt.strftime('%d/%m/%Y')

    return datos_historicos.to_dict(orient='records')

# Ubicom 

def histUbicom(file1, file2):
    # Leer el archivo de reporte diario
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Limpiar y convertir las columnas relevantes
    df1['Fecha'] = pd.to_datetime(df1['Fecha'], format='%Y-%m-%d').dt.date
    df1['Distancia recorrida'] = df1['Distancia recorrida'].astype(float)
    df1['Número de excesos de velocidad'] = df1['Número de excesos de velocidad'].astype(int)

    df2['Fecha'] = pd.to_datetime(df2['Fecha'], format='%Y-%m-%d').dt.date
    df2['Número'] = df2['Número'].astype(int)

    # Agrupar por fecha para sumar los kilómetros recorridos, contar los excesos de velocidad y los desplazamientos
    reporte_diario = df1.groupby('Fecha').agg({
        'Distancia recorrida': 'sum',
        'Número de excesos de velocidad': 'sum'
    }).rename(columns={
        'Distancia recorrida': 'km_recorridos',
        'Número de excesos de velocidad': 'num_excesos'
    }).reset_index()

    desplazamientos_diarios = df2.groupby('Fecha').agg({
        'Número': 'sum'
    }).rename(columns={'Número': 'num_desplazamientos'}).reset_index()

    # Unir los dataframes
    reporte_completo = pd.merge(reporte_diario, desplazamientos_diarios, on='Fecha', how='left')

    # Añadir la columna fija 'placa'
    reporte_completo['placa'] = 'FNM236'

    # Añadir columnas 'preoperacional' y 'día trabajado'
    reporte_completo['preoperacional'] = 1
    reporte_completo['dia_trabajado'] = 1

    # Si los km recorridos son igual a 0, entonces el número de desplazamientos también debe serlo y 'día trabajado' debe ser 0
    reporte_completo.loc[reporte_completo['km_recorridos'] == 0, 'num_desplazamientos'] = 0
    reporte_completo.loc[reporte_completo['km_recorridos'] == 0, 'dia_trabajado'] = 0
    reporte_completo.loc[reporte_completo['km_recorridos'] == 0, 'preoperacional'] = None

    # Seleccionar y renombrar columnas
    reporte_completo = reporte_completo[['placa', 'Fecha', 'km_recorridos', 'num_excesos', 'num_desplazamientos', 'preoperacional', 'dia_trabajado']]
    reporte_completo = reporte_completo.rename(columns={'Fecha': 'fecha'})

    # Convertir la columna 'fecha' a datetime
    reporte_completo['fecha'] = pd.to_datetime(reporte_completo['fecha'])

    # Convertir la fecha a formato dd/mm/yyyy
    reporte_completo['fecha'] = reporte_completo['fecha'].dt.strftime('%d/%m/%Y')

    return reporte_completo.to_dict(orient='records')

# Securitrac

def histSecuritrac(file):

    # Leer el archivo de Excel
    df = pd.read_excel(file)

    # Convertir la columna FECHAGPS a datetime
    df['FECHAGPS'] = pd.to_datetime(df['FECHAGPS'])

    # Extraer la fecha sin la hora
    df['Fecha'] = df['FECHAGPS'].dt.date

    # Calcular los kilómetros recorridos por día y placa, tomando el valor máximo
    km_recorridos = df.groupby(['NROMOVIL', 'Fecha'])['KILOMETROS'].max().reset_index()

    # Contar los excesos de velocidad por día y placa
    excesos_velocidad = df[df['EVENTO'] == 'Exc. Velocidad'].groupby(['NROMOVIL', 'Fecha']).size().reset_index(name='num_excesos')

    # Contar los desplazamientos por día y placa
    num_desplazamientos = df.groupby(['NROMOVIL', 'Fecha']).size().reset_index(name='num_desplazamientos')

    # Unir los datos de kilómetros recorridos, excesos de velocidad y desplazamientos
    datos_historicos = km_recorridos.merge(excesos_velocidad, on=['NROMOVIL', 'Fecha'], how='left').fillna(0)
    datos_historicos = datos_historicos.merge(num_desplazamientos, on=['NROMOVIL', 'Fecha'], how='left').fillna(0)

    # Añadir la columna fija 'placa'
    datos_historicos['placa'] = datos_historicos['NROMOVIL']

    # Añadir la columna día trabajado (1 si km_recorridos > 0, si no 0)
    datos_historicos['dia_trabajado'] = datos_historicos['KILOMETROS'].apply(lambda x: 1 if x > 0 else 0)

    # Añadir la columna preoperacional (1 si dia_trabajado es 1, None de lo contrario)
    datos_historicos['preoperacional'] = datos_historicos['dia_trabajado'].apply(lambda x: 1 if x == 1 else None)

    # Seleccionar y renombrar columnas necesarias
    datos_historicos = datos_historicos[['placa', 'Fecha', 'KILOMETROS', 'num_excesos', 'num_desplazamientos', 'preoperacional', 'dia_trabajado']]
    datos_historicos = datos_historicos.rename(columns={'Fecha': 'fecha', 'KILOMETROS': 'km_recorridos'})

    # Convertir la columna fecha a formato dd/mm/yyyy
    datos_historicos['fecha'] = pd.to_datetime(datos_historicos['fecha']).dt.strftime('%d/%m/%Y')

    return datos_historicos.to_dict(orient='records')

# Wialon

def histWialon(file_path):
    # Leer el archivo de Excel y obtener las hojas
    xls = pd.ExcelFile(file_path)
    
    # Leer la hoja 'Statistics' para obtener la placa
    df_stats = pd.read_excel(xls, 'Statistics')
    placa = df_stats.columns[1]  # Celda B1 en la hoja "Statistics"
    
    # Leer la hoja 'Excesos de velocidad' para contar los excesos por día
    df_excesos = pd.read_excel(xls, 'Excesos de velocidad')
    df_excesos['Comienzo'] = pd.to_datetime(df_excesos['Comienzo'])
    df_excesos['fecha'] = df_excesos['Comienzo'].dt.date
    num_excesos = df_excesos.groupby('fecha').size().reset_index(name='num_excesos')
    
    # Leer la hoja 'Cronología' para contar los desplazamientos por día
    df_cronologia = pd.read_excel(xls, 'Cronología')
    df_cronologia['Comienzo'] = pd.to_datetime(df_cronologia['Comienzo'])
    df_cronologia['fecha'] = df_cronologia['Comienzo'].dt.date
    num_desplazamientos = df_cronologia[df_cronologia['Tipo'] == 'Trip'].groupby('fecha').size().reset_index(name='num_desplazamientos')
    
    # Leer la hoja 'Calles visitadas' para sumar el kilometraje por día
    df_calles = pd.read_excel(xls, 'Calles visitadas')
    df_calles['Comienzo'] = pd.to_datetime(df_calles['Comienzo'])
    df_calles['fecha'] = df_calles['Comienzo'].dt.date
    df_calles['Kilometraje'] = round(df_calles['Kilometraje'].astype(str).str.replace(' km', '').astype(float), 2)
    km_recorridos = df_calles.groupby('fecha')['Kilometraje'].sum().reset_index(name='km_recorridos')
    
    # Combinar los datos de excesos, desplazamientos y kilometraje
    datos_historicos = pd.merge(num_excesos, num_desplazamientos, on='fecha', how='outer').fillna(0)
    datos_historicos = pd.merge(datos_historicos, km_recorridos, on='fecha', how='outer').fillna(0)
    
    # Añadir las columnas adicionales
    datos_historicos['placa'] = placa
    datos_historicos['dia_trabajado'] = (datos_historicos['num_desplazamientos'] > 0).astype(int)
    datos_historicos['preoperacional'] = datos_historicos['dia_trabajado'].apply(lambda x: 1 if x == 1 else None)
    
    # Seleccionar y renombrar columnas
    datos_historicos = datos_historicos[['placa', 'fecha', 'km_recorridos', 'num_excesos', 'num_desplazamientos', 'preoperacional', 'dia_trabajado']]
    datos_historicos.rename(columns={'fecha': 'Fecha'}, inplace=True)
    
    # Convertir la fecha a formato dd/mm/yyyy
    datos_historicos['Fecha'] = pd.to_datetime(datos_historicos['Fecha']).dt.strftime('%d/%m/%Y')
    
   
    
    return datos_historicos.to_dict(orient='records')

# Actualizar históricos

def agregar_datos_historicos(file_seguimiento, file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac, file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
    # Leer el archivo de seguimiento existente
    df_seguimiento = pd.read_excel(file_seguimiento)
    
    # Ejecutar las funciones de historial y recopilar los datos
    datos_ubicar = histUbicar(file_ubicar1, file_ubicar2)
    datos_MDVR = histMDVR(file_mdvr1, file_mdvr2)
    datos_ubicom = histUbicom(file_ubicom1, file_ubicom2)
    datos_securitrac = histSecuritrac(file_securitrac)
    datos_wialon1 = histWialon(file_wialon1)
    datos_wialon2 = histWialon(file_wialon2)
    datos_wialon3 = histWialon(file_wialon3)
    datos_ituran = histIturan(file_ituran1, file_ituran2)
    
    # Concatenar todos los datos históricos
    data_historica = datos_ubicar + datos_MDVR + datos_ubicom + datos_securitrac + datos_wialon1 + datos_wialon2 + datos_wialon3 + datos_ituran
    
    # Convertir los datos extraídos a un DataFrame
    df_hist = pd.DataFrame(data_historica)
    
    # Corregir las fechas NaN copiando el valor de la columna 'Fecha'
    df_hist['fecha'] = df_hist.apply(lambda row: row['Fecha'] if pd.isna(row['fecha']) else row['fecha'], axis=1)

    # Asegurar que la fecha esté en el formato correcto y tenga solo día y mes
    df_hist['fecha'] = pd.to_datetime(df_hist['fecha'], format='%d/%m/%Y').dt.strftime('%d/%m')
    
    print(f"Contenido de df_hist:")
    print(df_hist.head())
    
    # Rellenar los datos en el DataFrame existente con el formato deseado
    for index, row in df_hist.iterrows():
        fecha = row['fecha']
        placa = row['placa']
        
        if fecha in df_seguimiento.columns:
            if placa in df_seguimiento['PLACA'].values:
                df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Nº Excesos'), fecha] = row['num_excesos']
                df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Nº Desplazamiento'), fecha] = row['num_desplazamientos']
                df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Día Trabajado'), fecha] = row['dia_trabajado']
                df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Preoperacional'), fecha] = row['preoperacional']
                df_seguimiento.loc[(df_seguimiento['PLACA'] == placa) & (df_seguimiento['SEGUIMIENTO'] == 'Km recorridos'), fecha] = row['km_recorridos']
    
    # Guardar el archivo de seguimiento actualizado
    with pd.ExcelWriter(file_seguimiento, engine='openpyxl') as writer:
        df_seguimiento.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        worksheet.freeze_panes = worksheet['C2']  # Congela la primera fila y las dos primeras columnas

        worksheet.freeze_panes = worksheet['C2']  # Congela la primera fila y las dos primeras columnas

# Para correr esta parte de infractores del script necesitamos file_ituran2, fileMDVR2, file_ubicar2, file_Wailon (1, 2, 3), file_securitrac. 
# Nota: Todos estos archivos corresponden a los de excesos de velocidad. 

def infracIturan(file1):
    # Leer el archivo de Excel y obtener las hojas
    df = pd.read_excel(file1)
    
    # Filtrar la hoja que contiene la información relevante
    df_infracciones = df[['V_NICK_NAME', 'EVENT_DURATION_SEC', 'EVENT_DISTANCE', 'TOP_SPEED', 'VEHICLE_GROUP', 'ADDRESS', 'DRIVER_NAME', 'EVENT_START_DAY_TIME']]

    # Renombrar las columnas según lo solicitado
    df_infracciones.columns = ['PLACA', 'TIEMPO DE EXCESO', 'DURACIÓN EN KM DE EXCESO', 'VELOCIDAD MÁXIMA', 'PROYECTO', 'RUTA DE EXCESO', 'CONDUCTOR', 'FECHA']

    # Mantener la hora en la columna FECHA y convertirla al formato dd/mm/yyyy HH:MM:SS
    df_infracciones['FECHA'] = pd.to_datetime(df_infracciones['FECHA']).dt.strftime('%d/%m/%Y %H:%M:%S')

    # Filtrar los registros relevantes
    df_infracciones = df_infracciones[df_infracciones['DURACIÓN EN KM DE EXCESO'] > 0]

    num_registros = len(df_infracciones)
    print(f"Número de registros: {num_registros}")

    # Convertir el DataFrame en un diccionario
    datos_infracciones = df_infracciones.to_dict(orient='records')

    return datos_infracciones


#MDVR

# Función para convertir a segundos,

def convert_to_seconds(duration_str):
    
    parts = duration_str.split()
    minutes = 0
    seconds = 0
    for part in parts:
        if 'min' in part:
            minutes = int(part.replace('min', ''))
        elif 's' in part:
            seconds = int(part.replace('s', ''))
    return minutes * 60 + seconds

def infracMDVR(file1):
    # Leer el archivo Excel ignorando las primeras dos filas y la última fila
    df = pd.read_excel(file1, skiprows=2).iloc[:-1]
    
    # Extraer la placa del vehículo de la celda B1
    placa = pd.read_excel(file1).columns[1].replace('-', '')
    
    # Convertir la columna 'Comienzo' a datetime y extraer solo la fecha y hora
    df['Comienzo'] = pd.to_datetime(df['Comienzo'], dayfirst= True)
    df['Fecha'] = df['Comienzo'].dt.strftime('%d/%m/%Y %H:%M:%S')
    
    # Dividir la columna 'Posición' en 'Latitud' y 'Longitud'
    df[['Latitud', 'Longitud']] = df['Posición'].str.split(',', expand=True)
    df['Latitud'] = df['Latitud'].astype(float)
    df['Longitud'] = df['Longitud'].astype(float)
    
    # Convertir el tiempo de exceso a segundos
    df['Duración exceso de velocidad'] = df['Duración exceso de velocidad'].apply(convert_to_seconds)
    
    # Crear el diccionario con el formato requerido
    registros = []
    for index, row in df.iterrows():
        registro = {
            'PLACA': placa,
            'TIEMPO DE EXCESO': row['Duración exceso de velocidad'],
            'DURACIÓN EN KM DE EXCESO': '',
            'VELOCIDAD MÁXIMA': float(row['Velocidad máxima'].replace('kph', '')),
            'PROYECTO': '',
            'RUTA DE EXCESO': f"{row['Latitud']}, {row['Longitud']}",  # Guardar solo las coordenadas
            'CONDUCTOR': '',
            'FECHA': row['Fecha']
        }
        registros.append(registro)
    
    print(f"Número total de registros: {len(registros)}")
    return registros

# Ubicar

def infracUbicar(file1):
    # Leer el archivo de Excel y obtener las hojas
    df = pd.read_excel(file1, skiprows=2).iloc[:-1]  # Ignorar las primeras 4 filas y la última fila

    # Extraer la placa del vehículo de la celda B1
    placa = "JYT620"  # Reemplazar con la extracción correcta si se requiere

    # Convertir la columna 'Comienzo' a datetime y extraer solo la fecha y hora
    df['Comienzo'] = pd.to_datetime(df['Comienzo'], dayfirst= True)
    df['Fecha'] = df['Comienzo'].dt.strftime('%d/%m/%Y %H:%M:%S')
    
    # Convertir la columna 'Duración' a segundos
    def convert_duration_to_seconds(duration):
        parts = duration.split(' ')
        total_seconds = 0
        for part in parts:
            if 'h' in part:
                total_seconds += int(part.replace('h', '')) * 3600
            elif 'min' in part:
                total_seconds += int(part.replace('min', '')) * 60
            elif 's' in part:
                total_seconds += int(part.replace('s', ''))
        return total_seconds

    df['Tiempo de Exceso'] = df['Duración'].apply(convert_duration_to_seconds)
    
    # Crear el diccionario con el formato requerido
    registros = []
    for index, row in df.iterrows():
        registros.append({
            'PLACA': placa,
            'TIEMPO DE EXCESO': row['Tiempo de Exceso'],
            'DURACIÓN EN KM DE EXCESO': None,
            'VELOCIDAD MÁXIMA': float(row['Velocidad máxima'].replace(' kph', '')),
            'PROYECTO': None,
            'CONDUCTOR': None,
            'RUTA DE EXCESO': row['Posición'],
            'FECHA': row['Fecha']
        })

    # Imprimir el número de registros
    print(f'Número de registros: {len(registros)}')
    
    # Retornar el diccionario de registros
    return registros

# Securitrac 

def infracSecuritrac(file):
    # Leer el archivo de Excel
    df = pd.read_excel(file)

    # Filtrar solo las filas con "Exc. Velocidad"
    df = df[df['EVENTO'] == 'Exc. Velocidad']

    # Crear el diccionario con el formato requerido
    registros = []
    for index, row in df.iterrows():
        fecha_formateada = pd.to_datetime(row['FECHAGPS']).strftime('%d/%m/%Y %H:%M:%S')
        registro = {
            'PLACA': row['NROMOVIL'],
            'TIEMPO DE EXCESO': None,
            'DURACIÓN EN KM DE EXCESO': None,
            'VELOCIDAD MÁXIMA': row['VELOCIDAD'],
            'RUTA DE EXCESO': row['POSICION'],
            'PROYECTO': None,
            'CONDUCTOR': None,
            'FECHA': fecha_formateada,
        }
        registros.append(registro)

    print(f"Total de registros generados: {len(registros)}")
    print(registros[:5]) # Muestra los primeros 5 registros para verificar

    return registros

# Wialon

def infracWialon(file1):
    xls = pd.ExcelFile(file1)

    # Leer la hoja 'Statistics' para obtener la placa
    df_stats = pd.read_excel(xls, 'Statistics')
    placa = df_stats.columns[1]  # Celda B1 en la hoja "Statistics"

    # Leer la hoja 'Excesos de velocidad' para obtener la información necesaria
    df_excesos = pd.read_excel(xls, 'Excesos de velocidad')

    # Convertir la columna 'Comienzo' a datetime y extraer solo la fecha y hora
    df_excesos['Comienzo'] = pd.to_datetime(df_excesos['Comienzo'])
    df_excesos['Fecha'] = df_excesos['Comienzo'].dt.strftime('%d/%m/%Y %H:%M:%S')

    # Convertir la columna 'Duración' a segundos
    def convert_duration_to_seconds(duration_str):
        parts = duration_str.split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        else:
            return int(parts[0])

    df_excesos['Duración en segundos'] = df_excesos['Duración'].apply(convert_duration_to_seconds)

    # Crear el diccionario con el formato requerido
    registros = []
    for index, row in df_excesos.iterrows():
        registro = {
            'PLACA': placa,
            'TIEMPO DE EXCESO': row['Duración en segundos'],
            'DURACIÓN EN KM DE EXCESO': '',  # No disponible en este documento
            'VELOCIDAD MÁXIMA': row['Velocidad máxima'],
            'PROYECTO': '',  # No disponible en este documento
            'CONDUCTOR': '',  # No disponible en este documento
            'RUTA DE EXCESO': row['Localización'],
            'FECHA': row['Fecha']
        }
        registros.append(registro)

    print(f"Total de registros extraídos: {len(registros)}")
    return registros

# Ejecutar todos

def infracTodos(fileIturan, fileMDVR, fileUbicar, fileWialon1, fileWialon2, fileWialon3, fileSecuritrac):
    # Ejecutar cada función infrac y obtener los resultados
    registros_ituran = infracIturan(fileIturan)
    registros_mdvr = infracMDVR(fileMDVR)
    registros_ubicar = infracUbicar(fileUbicar)
    registros_wialon = infracWialon(fileWialon1)
    registros_wialon2 = infracWialon(fileWialon2)
    registros_wialon3 = infracWialon(fileWialon3)
    registros_securitrac = infracSecuritrac(fileSecuritrac)

    # Combinar todos los resultados en una sola lista
    todos_registros = (
        registros_ituran + registros_mdvr + registros_ubicar + registros_wialon + registros_wialon2 + registros_wialon3 + registros_securitrac
    )

    print(f"Total de registros combinados: {len(todos_registros)}")
    return todos_registros


# Esta es la función que crea por PRIMERA vez la hoja de infractores.

def crearInfractores(file_seguimiento, file_Ituran, file_MDVR, file_Ubicar, file_Wialon, file_Securitrac):
    # Obtener todas las infracciones combinadas
    todos_registros = infracTodos(file_Ituran, file_MDVR, file_Ubicar, file_Wialon, file_Securitrac)
    df_infractores = pd.DataFrame(todos_registros)

    # Convertir la columna 'FECHA' a datetime y luego a string con el formato correcto
    df_infractores['FECHA'] = pd.to_datetime(df_infractores['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')

    df_infractores = df_infractores[(df_infractores['VELOCIDAD MÁXIMA'] > 80) & (df_infractores['TIEMPO DE EXCESO'] > 20)]

    # Cargar el archivo existente y añadir una nueva hoja
    with pd.ExcelWriter(file_seguimiento, engine='openpyxl', mode='a') as writer:
        df_infractores.to_excel(writer, sheet_name='Infractores', index=False)

# Esta función es la que llena la tabla de infractores en la base de datos 'vehiculos' con la información histórica. 

###### ESTO ES PARA ACTUALIZAR LAS BASES DE DATOS EN MySQL ########

def actualizarInfractoresSQL(file_Ituran, file_MDVR, file_Ubicar, file_Wialon, file_Securitrac): 
        # Obtener todas las infracciones combinadas
    todos_registros = infracTodos(file_Ituran, file_MDVR, file_Ubicar, file_Wialon, file_Securitrac)
    df_infractores = pd.DataFrame(todos_registros)

    # Convertir la columna 'FECHA' a datetime y luego a string con el formato correcto
    df_infractores['FECHA'] = pd.to_datetime(df_infractores['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')

    df_infractores = df_infractores[(df_infractores['VELOCIDAD MÁXIMA'] > 80) & (df_infractores['TIEMPO DE EXCESO'] > 20)]

    # Conexión

    user = 'root'
    password = '123456678'  
    host = 'localhost'
    port = '3306'
    schema = 'vehiculos'

    engine = create_engine(f'mysql+mysqlconnector://{user}:{password}@{host}:{port}/{schema}')

     # Insertar los datos en la tabla seguimiento
    df_infractores.to_sql('infractores', con=engine, if_exists='append', index=False)

    print("Datos insertados correctamente en la tabla seguimiento.")

    return(df_infractores)

# Estas funciones solo son pequeñas modificaciones de las de extracción de los datos históricos para llenar la tabla de seguimiento con los datos históricos

def sqlIturan(file1, file2):
    # Leer los archivos Excel
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Convertir las fechas a un formato de fecha datetime
    df1['DT_DRIVE_DATE'] = pd.to_datetime(df1['DT_DRIVE_DATE']).dt.date
    df2['EVENT_START_DAY_TIME'] = pd.to_datetime(df2['EVENT_START_DAY_TIME']).dt.date

    # Crear un DataFrame para almacenar los resultados
    results = []

    # Obtener las placas únicas
    placas = df1['V_NICK_NAME'].unique()

    for placa in placas:
        # Filtrar los datos por placa
        df1_placa = df1[df1['V_NICK_NAME'] == placa]
        df2_placa = df2[df2['V_NICK_NAME'] == placa]

        # Agrupar por fecha para contar los desplazamientos y sumar las distancias
        daily_stats_df1 = df1_placa.groupby('DT_DRIVE_DATE').agg({
            'TRIP_DISTANCE': 'sum',
            'DT_DRIVE_DATE': 'count'  # Esta cuenta el número de desplazamientos
        }).rename(columns={'DT_DRIVE_DATE': 'num_desplazamientos', 'TRIP_DISTANCE': 'km_recorridos'}).reset_index()

        # Agrupar por fecha para contar los excesos
        daily_stats_df2 = df2_placa.groupby('EVENT_START_DAY_TIME').size().reset_index(name='num_excesos')

        # Unir ambos DataFrames por fecha
        daily_stats = pd.merge(daily_stats_df1, daily_stats_df2, left_on='DT_DRIVE_DATE', right_on='EVENT_START_DAY_TIME', how='left').fillna(0)

        current_datetime = datetime.now()
        current_time = current_datetime.strftime("%H:%M:%S")

        for _, row in daily_stats.iterrows():
            fecha_formateada = row['DT_DRIVE_DATE'].strftime('%d/%m/%Y')
            datos_extraidos = {
                'placa': placa,
                'fecha': fecha_formateada + ' ' + current_time,
                'km_recorridos': round(row['km_recorridos'], 2),
                'num_desplazamientos': row['num_desplazamientos'],
                'dia_trabajado': 1 if row['km_recorridos'] > 0 else 0,
                'preoperacional': 1 if row['km_recorridos'] > 0 else None,
                'num_excesos': int(row['num_excesos']),
                'proveedor': 'Ituran'
            }
            results.append(datos_extraidos)

    return results

def sqlMDVR(file1, file2):
    # Leer el archivo de desplazamientos
    df1 = pd.read_excel(file1, engine='xlrd', header=2, skipfooter=2)
    df1['Fecha'] = pd.to_datetime(df1['Fecha'], format='%d/%m/%Y %H:%M:%S').dt.date
    
    # Limpiar y convertir 'Longitud de ruta' a float
    df1['Longitud de ruta'] = df1['Longitud de ruta'].str.replace(' Km', '').astype(float)
    
    # Contar desplazamientos y sumar kilómetros por día
    desplazamientos_km = df1.groupby('Fecha').agg(
        num_desplazamientos=('Fecha', 'size'),
        km_recorridos=('Longitud de ruta', 'sum')
    ).reset_index()
    
    
    file_path = file2
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Extraer los datos de la hoja y convertirlos en un DataFrame
    data = sheet.values
    next(data) 
    next(data) 

    next(data) 
    

    cols = ["Comienzo", "Fin", "Duración exceso de velocidad", "Velocidad máxima", "Velocidad media", "Posición"]
    data = list(data)
    df2 = pd.DataFrame(data, columns=cols)
    
    # Quitar la última fila
    df2 = df2.iloc[:-1]
    
    # Convertir la columna 'Comienzo' a datetime
    df2['Comienzo'] = pd.to_datetime(df2['Comienzo'], format='%d/%m/%Y %H:%M:%S').dt.date
    
    # Contar excesos de velocidad por día
    excesos_velocidad = df2.groupby('Comienzo').size().reset_index(name='num_excesos')
    
    # Unir los datos de desplazamientos y excesos de velocidad
    datos_historicos = pd.merge(desplazamientos_km, excesos_velocidad, left_on='Fecha', right_on='Comienzo', how='left').fillna(0)
    
    # Añadir la columna fija 'placa'
    datos_historicos['placa'] = 'KSZ298'
    
    # Añadir las columnas faltantes
    datos_historicos['dia_trabajado'] = (datos_historicos['km_recorridos'] > 0).astype(int)
    datos_historicos['preoperacional'] = datos_historicos['dia_trabajado'].replace(0, '-')
    
    # Seleccionar y renombrar columnas
    datos_historicos = datos_historicos[['placa', 'Fecha', 'km_recorridos', 'num_desplazamientos', 'dia_trabajado', 'preoperacional', 'num_excesos']]
    datos_historicos = datos_historicos.rename(columns={'Fecha': 'fecha'})
    
    # Convertir la columna 'fecha' a datetime
    datos_historicos['fecha'] = pd.to_datetime(datos_historicos['fecha'])

    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")
    
    # Convertir la fecha a formato dd/mm/yyyy
    datos_historicos['fecha'] = datos_historicos['fecha'].dt.strftime('%d/%m/%Y') + ' ' + current_time

    # Asignar una nueva columna 'proveedor' con todas las entradas como 'Wialon'
    datos_historicos['proveedor'] = 'MDVR'
    
    return datos_historicos.to_dict(orient='records')

def sqlUbicar(file1, file2):
    # Leer el archivo de desplazamientos
    df1 = pd.read_excel(file1, engine='xlrd', header=2, skipfooter=2)
    df1['Fecha'] = df1['Fecha'].apply(lambda x: x.replace('-', '/'))
    df1['Fecha'] = pd.to_datetime(df1['Fecha'], format="%d/%m/%Y %H:%M:%S").dt.date

    # Limpiar y convertir 'Longitud de ruta' a float
    df1['Longitud de ruta'] = df1['Longitud de ruta'].str.replace(' Km', '').astype(float)

    # Contar desplazamientos y sumar kilómetros por día
    desplazamientos_km = df1.groupby('Fecha').agg(
        num_desplazamientos=('Fecha', 'size'),
        km_recorridos=('Longitud de ruta', 'sum')
    ).reset_index()

    # Leer el archivo de excesos de velocidad
    workbook = openpyxl.load_workbook(file2)
    sheet = workbook.active

    # Extraer los datos de la hoja y convertirlos en un DataFrame
    data = sheet.values
    next(data) # Saltar la primera fila de datos (encabezados principales)
    next(data) # Saltar la segunda fila de datos (encabezados secundarios)
    next(data)

    cols = ["Comienzo", "Fin", "Duración exceso de velocidad", "Velocidad máxima", "Velocidad media", "Posición"]
    data = list(data)
    df2 = pd.DataFrame(data, columns=cols)

    # Quitar la última fila
    df2 = df2.iloc[:-1]

    # Convertir la columna 'Comienzo' a datetime
    df2['Comienzo'] = pd.to_datetime(df2['Comienzo'].str.replace('.', '/'), format="%d/%m/%Y %H:%M:%S", errors='coerce').dt.date

    # Contar excesos de velocidad por día
    excesos_velocidad = df2.groupby('Comienzo').size().reset_index(name='num_excesos')

    # Unir los datos de desplazamientos y excesos de velocidad
    datos_historicos = pd.merge(desplazamientos_km, excesos_velocidad, left_on='Fecha', right_on='Comienzo', how='left').fillna(0)

    # Añadir la columna fija 'placa'
    datos_historicos['placa'] = 'JYT682'

    # Añadir las columnas faltantes
    datos_historicos['dia_trabajado'] = (datos_historicos['km_recorridos'] > 0).astype(int)
    datos_historicos['preoperacional'] = (datos_historicos['km_recorridos'] > 0).astype(int)
    datos_historicos.loc[datos_historicos['dia_trabajado'] == 0, 'preoperacional'] = '-'

    # Seleccionar y renombrar columnas
    datos_historicos = datos_historicos[['placa', 'Fecha', 'km_recorridos', 'num_desplazamientos', 'dia_trabajado', 'preoperacional', 'num_excesos']]
    datos_historicos = datos_historicos.rename(columns={'Fecha': 'fecha'})

    # Convertir la columna 'fecha' a datetime
    datos_historicos['fecha'] = pd.to_datetime(datos_historicos['fecha'])

    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    # Convertir la fecha a formato dd/mm/yyyy
    datos_historicos['fecha'] = datos_historicos['fecha'].dt.strftime('%d/%m/%Y') + ' ' + current_time

    # Asignar una nueva columna 'proveedor' con todas las entradas como 'Wialon'
    datos_historicos['proveedor'] = 'Ubicar'

    return datos_historicos.to_dict(orient='records')

def sqlUbicom(file1, file2):
    # Leer el archivo de reporte diario
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Limpiar y convertir las columnas relevantes
    df1['Fecha'] = pd.to_datetime(df1['Fecha'], format='%Y-%m-%d').dt.date
    df1['Distancia recorrida'] = df1['Distancia recorrida'].astype(float)
    df1['Número de excesos de velocidad'] = df1['Número de excesos de velocidad'].astype(int)

    df2['Fecha'] = pd.to_datetime(df2['Fecha'], format='%Y-%m-%d').dt.date
    df2['Número'] = df2['Número'].astype(int)

    # Agrupar por fecha para sumar los kilómetros recorridos, contar los excesos de velocidad y los desplazamientos
    reporte_diario = df1.groupby('Fecha').agg({
        'Distancia recorrida': 'sum',
        'Número de excesos de velocidad': 'sum'
    }).rename(columns={
        'Distancia recorrida': 'km_recorridos',
        'Número de excesos de velocidad': 'num_excesos'
    }).reset_index()

    desplazamientos_diarios = df2.groupby('Fecha').agg({
        'Número': 'sum'
    }).rename(columns={'Número': 'num_desplazamientos'}).reset_index()

    # Unir los dataframes
    reporte_completo = pd.merge(reporte_diario, desplazamientos_diarios, on='Fecha', how='left')

    # Añadir la columna fija 'placa'
    reporte_completo['placa'] = 'FNM236'

    # Añadir columnas 'preoperacional' y 'día trabajado'
    reporte_completo['preoperacional'] = 1
    reporte_completo['dia_trabajado'] = 1

    # Si los km recorridos son igual a 0, entonces el número de desplazamientos también debe serlo y 'día trabajado' debe ser 0
    reporte_completo.loc[reporte_completo['km_recorridos'] == 0, 'num_desplazamientos'] = 0
    reporte_completo.loc[reporte_completo['km_recorridos'] == 0, 'dia_trabajado'] = 0
    reporte_completo.loc[reporte_completo['km_recorridos'] == 0, 'preoperacional'] = '-'

    # Seleccionar y renombrar columnas
    reporte_completo = reporte_completo[['placa', 'Fecha', 'km_recorridos', 'num_excesos', 'num_desplazamientos', 'preoperacional', 'dia_trabajado']]
    reporte_completo = reporte_completo.rename(columns={'Fecha': 'fecha'})

    # Convertir la columna 'fecha' a datetime
    reporte_completo['fecha'] = pd.to_datetime(reporte_completo['fecha'])

    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    # Convertir la fecha a formato dd/mm/yyyy
    reporte_completo['fecha'] = reporte_completo['fecha'].dt.strftime('%d/%m/%Y') + ' ' + current_time

    # Asignar una nueva columna 'proveedor' con todas las entradas como 'Wialon'
    reporte_completo['proveedor'] = 'Ubicom'

    return reporte_completo.to_dict(orient='records')

def sqlSecuritrac(file):
    # Leer el archivo de Excel
    df = pd.read_excel(file)

    # Convertir la columna FECHAGPS a datetime
    df['FECHAGPS'] = pd.to_datetime(df['FECHAGPS'])

    # Extraer la fecha sin la hora
    df['Fecha'] = df['FECHAGPS'].dt.date

    # Calcular los kilómetros recorridos por día y placa, tomando el valor máximo
    km_recorridos = df.groupby(['NROMOVIL', 'Fecha'])['KILOMETROS'].max().reset_index()

    # Contar los excesos de velocidad por día y placa
    excesos_velocidad = df[df['EVENTO'] == 'Exc. Velocidad'].groupby(['NROMOVIL', 'Fecha']).size().reset_index(name='num_excesos')

    # Contar los desplazamientos por día y placa
    num_desplazamientos = df.groupby(['NROMOVIL', 'Fecha']).size().reset_index(name='num_desplazamientos')

    # Unir los datos de kilómetros recorridos, excesos de velocidad y desplazamientos
    datos_historicos = km_recorridos.merge(excesos_velocidad, on=['NROMOVIL', 'Fecha'], how='left').fillna(0)
    datos_historicos = datos_historicos.merge(num_desplazamientos, on=['NROMOVIL', 'Fecha'], how='left').fillna(0)

    # Añadir la columna fija 'placa'
    datos_historicos['placa'] = datos_historicos['NROMOVIL']

    # Añadir la columna día trabajado (1 si km_recorridos > 0, si no 0)
    datos_historicos['dia_trabajado'] = datos_historicos['KILOMETROS'].apply(lambda x: 1 if x > 0 else 0)

    # Añadir la columna preoperacional (1 si dia_trabajado es 1, None de lo contrario)
    datos_historicos['preoperacional'] = datos_historicos['dia_trabajado'].apply(lambda x: 1 if x == 1 else '-')

    # Seleccionar y renombrar columnas necesarias
    datos_historicos = datos_historicos[['placa', 'Fecha', 'KILOMETROS', 'num_excesos', 'num_desplazamientos', 'preoperacional', 'dia_trabajado']]
    datos_historicos = datos_historicos.rename(columns={'Fecha': 'fecha', 'KILOMETROS': 'km_recorridos'})
    
    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    # Convertir la columna fecha a formato dd/mm/yyyy
    datos_historicos['fecha'] = pd.to_datetime(datos_historicos['fecha']).dt.strftime('%d/%m/%Y') + ' ' + current_time

    # Asignar una nueva columna 'proveedor' con todas las entradas como 'Wialon'
    datos_historicos['proveedor'] = 'Securitrac'

    return datos_historicos.to_dict(orient='records')

def sqlWialon(file_path):
    # Leer el archivo de Excel y obtener las hojas
    xls = pd.ExcelFile(file_path)
    
    # Leer la hoja 'Statistics' para obtener la placa
    df_stats = pd.read_excel(xls, 'Statistics')
    placa = df_stats.columns[1]  # Celda B1 en la hoja "Statistics"
    
    # Leer la hoja 'Excesos de velocidad' para contar los excesos por día
    df_excesos = pd.read_excel(xls, 'Excesos de velocidad')
    df_excesos['Comienzo'] = pd.to_datetime(df_excesos['Comienzo'])
    df_excesos['fecha'] = df_excesos['Comienzo'].dt.date
    num_excesos = df_excesos.groupby('fecha').size().reset_index(name='num_excesos')
    
    # Leer la hoja 'Cronología' para contar los desplazamientos por día
    df_cronologia = pd.read_excel(xls, 'Cronología')
    df_cronologia['Comienzo'] = pd.to_datetime(df_cronologia['Comienzo'])
    df_cronologia['fecha'] = df_cronologia['Comienzo'].dt.date
    num_desplazamientos = df_cronologia[df_cronologia['Tipo'] == 'Trip'].groupby('fecha').size().reset_index(name='num_desplazamientos')
    
    # Leer la hoja 'Calles visitadas' para sumar el kilometraje por día
    df_calles = pd.read_excel(xls, 'Calles visitadas')
    df_calles['Comienzo'] = pd.to_datetime(df_calles['Comienzo'])
    df_calles['fecha'] = df_calles['Comienzo'].dt.date
    df_calles['Kilometraje'] = round(df_calles['Kilometraje'].astype(str).str.replace(' km', '').astype(float), 2)
    km_recorridos = df_calles.groupby('fecha')['Kilometraje'].sum().reset_index(name='km_recorridos')
    
    # Combinar los datos de excesos, desplazamientos y kilometraje
    datos_historicos = pd.merge(num_excesos, num_desplazamientos, on='fecha', how='outer').fillna(0)
    datos_historicos = pd.merge(datos_historicos, km_recorridos, on='fecha', how='outer').fillna(0)
    
    # Añadir las columnas adicionales
    datos_historicos['placa'] = placa
    datos_historicos['dia_trabajado'] = (datos_historicos['num_desplazamientos'] > 0).astype(int)
    datos_historicos['preoperacional'] = datos_historicos['dia_trabajado'].apply(lambda x: 1 if x == 1 else '-')
    
    # Seleccionar y renombrar columnas
    datos_historicos = datos_historicos[['placa', 'fecha', 'km_recorridos', 'num_excesos', 'num_desplazamientos', 'preoperacional', 'dia_trabajado']]
    datos_historicos.rename(columns={'fecha': 'Fecha'}, inplace=True)

    current_datetime = datetime.now()
    current_time = current_datetime.strftime("%H:%M:%S")

    
    # Convertir la fecha a formato dd/mm/yyyy
    datos_historicos['Fecha'] = pd.to_datetime(datos_historicos['Fecha']).dt.strftime('%d/%m/%Y') + ' ' + current_time


    # Asignar una nueva columna 'proveedor' con todas las entradas como 'Wialon'
    datos_historicos['proveedor'] = 'Wialon'
    
    return datos_historicos.to_dict(orient='records')

def juntarDatosSQL(file_ituran1, file_ituran2, file_mdvr1, file_mdvr2, file_ubicar1, file_ubicar2, file_ubicom1, file_ubicom2, file_securitrac, file_wialon1, file_wialon2, file_wialon3):
 
    # Ejecutar las funciones y recopilar los datos
    datos_ubicar = sqlUbicar(file_ubicar1, file_ubicar2)
    datos_MDVR = sqlMDVR(file_mdvr1, file_mdvr2)
    datos_ubicom = sqlUbicom(file_ubicom1, file_ubicom2)
    datos_securitrac = sqlSecuritrac(file_securitrac)
    datos_wialon1 = sqlWialon(file_wialon1)
    datos_wialon2 = sqlWialon(file_wialon2)
    datos_wialon3 = sqlWialon(file_wialon3)
    datos_ituran = sqlIturan(file_ituran1, file_ituran2)
    
    # Concatenar todos los datos 
    dataHistorica = datos_ituran + datos_MDVR + datos_ubicar +  datos_ubicom + datos_securitrac + datos_wialon1 + datos_wialon2 + datos_wialon3
    # Convertir los datos extraídos a un DataFrame
    df_hist = pd.DataFrame(dataHistorica)
    df_hist['Fecha'] = df_hist['Fecha'].fillna(df_hist['fecha'])

    # Eliminar la columna 'fecha'
    df_hist.drop(columns=['fecha'], inplace=True)

    df_hist.rename(columns={
    'placa': 'placa',
    'km_recorridos': 'kmRecorridos',
    'num_excesos': 'numExcesos',
    'num_desplazamientos': 'numDesplazamientos',
    'dia_trabajado': 'diaTrabajado',
    'preoperacional': 'preoperacional',
    'proveedor': 'proveedor',
    'Fecha': 'fecha'  
    }, inplace=True)

    df_hist.fillna(0, inplace=True)

    return df_hist



#### INDICADORES ###

def crearDfHist(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                  file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
    
    data_historica = []
    
    data_historica.extend(histUbicom(file_ubicom1, file_ubicom2))
    data_historica.extend(histUbicar(file_ubicar1, file_ubicar2))
    data_historica.extend(histMDVR(file_mdvr1, file_mdvr2))
    data_historica.extend(histSecuritrac(file_securitrac))
    data_historica.extend(histWialon(file_wialon1))
    data_historica.extend(histWialon(file_wialon2))
    data_historica.extend(histWialon(file_wialon3))
    data_historica.extend(histIturan(file_ituran1, file_ituran2))
    
    df_hist = pd.DataFrame(data_historica)
    

# Corregir las fechas NaN copiando el valor de la columna 'Fecha'

    df_hist['fecha'] = df_hist.apply(lambda row: row['Fecha'] if pd.isna(row['fecha']) else row['fecha'], axis=1)

    df_hist['fecha'] = pd.to_datetime(df_hist['fecha'], format='%d/%m/%Y')

    return df_hist
    
def crear_df_diario(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                  file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
    
    df_hist =  crearDfHist(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                  file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2)
    # Asegurarse de que las columnas existan en df_hist y en el formato esperado
    expected_columns = ['placa', 'fecha', 'num_excesos', 'num_desplazamientos', 'dia_trabajado', 'preoperacional', 'km_recorridos']

    for col in expected_columns:
        if col not in df_hist.columns:
            df_hist[col] = None

    df_diario = df_hist.groupby('fecha').agg({
        'km_recorridos': 'sum',
        'num_excesos': 'sum',
        'num_desplazamientos': 'sum',
        'dia_trabajado': 'sum',
        'preoperacional': 'sum'
    }).reset_index()

    # Renombramos las columnas para que coincidan con el formato deseado
    df_diario.columns = ['FECHA', 'KILOMETROS RECORRIDOS', 'EXCESOS VELOCIDAD', 'DESPLAZAMIENTOS', 'DÍA TRABAJADO', 'PREOPERACIONAL']

    # Ordenar el DataFrame por la columna 'FECHA'
    df_diario = df_diario.sort_values(by='FECHA', key=lambda x: pd.to_datetime(x, format='%d/%m/%Y'))

    # Reiniciar el índice si es necesario
    df_diario = df_diario.reset_index(drop=True)

    return df_diario

def exportar_df_diario(file_seguimiento, file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                  file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2):
    
    df_diario = crear_df_diario(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,
                  file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2)
    # Convertir la columna 'FECHA' a datetime y luego formatear para quitar la hora
    df_diario['FECHA'] = pd.to_datetime(df_diario['FECHA'], format="%d/%m/%Y").dt.strftime('%d/%m/%Y')

    with pd.ExcelWriter(file_seguimiento, mode='a', engine='openpyxl') as writer:
        df_diario.to_excel(writer, sheet_name='Indicadores Totales', index=False)

    print(f"Agregado como hoja 'Indicadores Totales' en {file_seguimiento}")



def calcular_EJL(df_diario):
    # Asegurarse de que la columna 'FECHA' es de tipo datetime en el DataFrame original
    df_diario['FECHA'] = pd.to_datetime(df_diario['FECHA'], format="%d/%m/%Y")

    # Crear una columna para el mes y año
    df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')

    # Calcular los acumulados mensuales
    df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()

    # Crear un DataFrame con todos los meses del año
    all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
    df_all_months = pd.DataFrame(all_months, columns=['MES'])

    # Unir con el DataFrame de acumulados mensuales para asegurarse de que todos los meses estén presentes
    df_acumulados = pd.merge(df_all_months, df_acumulados, on='MES', how='left')

    # Rellenar los NaN con ceros y asegurarse de que el tipo de datos sea correcto
    df_acumulados.fillna(0, inplace=True)

    # Cambiar el formato del periodo a nombre de mes en español
    df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()

    # Calcular EJL
    df_EJL = pd.DataFrame({
        'MES': df_acumulados['MES'],
        'EJD': df_acumulados['EXCESOS VELOCIDAD'],
        'SDT': df_acumulados['DÍA TRABAJADO'],
        'EJL': (df_acumulados['EXCESOS VELOCIDAD'] / df_acumulados['DÍA TRABAJADO']) * 100
    })

    # Redondear a 2 decimales
    df_EJL = df_EJL.round(2)

    # Transponer el DataFrame
    df_EJL = df_EJL.set_index('MES').transpose()

    return df_EJL

def calcular_GVE(df_diario, df_hist):
    # Asegurarse de que la columna 'FECHA' es de tipo datetime en el DataFrame original
    df_diario['FECHA'] = pd.to_datetime(df_diario['FECHA'], format="%d/%m/%Y")

    # Crear una columna para el mes y año
    df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')

    # Calcular los acumulados mensuales
    df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()

    # Calcular el valor máximo de "DÍA TRABAJADO" por mes
    df_max_dia_trabajado = df_diario.groupby('MES')['DÍA TRABAJADO'].max().reset_index()

    # Crear un DataFrame con todos los meses del año
    all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
    df_all_months = pd.DataFrame(all_months, columns=['MES'])

    # Unir los DataFrames de acumulados mensuales y máximos de día trabajado para asegurarse de que todos los meses estén presentes
    df_acumulados = pd.merge(df_all_months, df_acumulados, on='MES', how='left')
    df_max_dia_trabajado = pd.merge(df_all_months, df_max_dia_trabajado, on='MES', how='left')

    # Rellenar los NaN con ceros y asegurarse de que el tipo de datos sea correcto
    df_acumulados.fillna(0, inplace=True)
    df_max_dia_trabajado.fillna(0, inplace=True)

    # Cambiar el formato del periodo a nombre de mes en español
    df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()
    df_max_dia_trabajado['MES'] = df_max_dia_trabajado['MES'].dt.strftime('%B').str.capitalize()

    # Calcular VIP y VLD
    vip_value = len(df_hist['placa'].unique())  # VIP es constante para todos los meses
    vld_values = df_max_dia_trabajado['DÍA TRABAJADO'].tolist()  # VLD es el valor máximo de DÍA TRABAJADO para cada mes

    # Calcular GVE
    gve_values = [(vip_value / vld) * 100 if vld != 0 else 0 for vld in vld_values]

    # Crear el DataFrame de GVE
    df_GVE = pd.DataFrame({
        'MES': df_acumulados['MES'],
        'VIP': [vip_value] * len(df_acumulados),
        'VLD': vld_values,
        'GVE': gve_values
    })

    # Redondear a 2 decimales
    df_GVE = df_GVE.round(2)

    # Transponer el DataFrame
    df_GVE = df_GVE.set_index('MES').transpose()

    return df_GVE

def calcular_ELVL(df_diario):
    # Asegurarse de que la columna 'FECHA' es de tipo datetime en el DataFrame original
    df_diario['FECHA'] = pd.to_datetime(df_diario['FECHA'], format="%d/%m/%Y")

    # Crear una columna para el mes y año
    df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')

    # Calcular los acumulados mensuales
    df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()

    # Crear un DataFrame con todos los meses del año
    all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
    df_all_months = pd.DataFrame(all_months, columns=['MES'])

    # Unir con el DataFrame de acumulados mensuales para asegurarse de que todos los meses estén presentes
    df_acumulados = pd.merge(df_all_months, df_acumulados, on='MES', how='left')

    # Rellenar los NaN con ceros y asegurarse de que el tipo de datos sea correcto
    df_acumulados.fillna(0, inplace=True)

    # Cambiar el formato del periodo a nombre de mes en español
    df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()

    # Calcular ELVL
    df_ELVL = pd.DataFrame({
        'MES': df_acumulados['MES'],
        'DLEV': df_acumulados['EXCESOS VELOCIDAD'],
        'TDL': df_acumulados['DESPLAZAMIENTOS'],
        'ELVL': (df_acumulados['EXCESOS VELOCIDAD'] / df_acumulados['DESPLAZAMIENTOS']) * 100
    })

    # Redondear a 2 decimales
    df_ELVL = df_ELVL.round(2)

    # Transponer el DataFrame
    df_ELVL = df_ELVL.set_index('MES').transpose()

    return df_ELVL

def calcular_IDP(df_diario):
    # Asegurarse de que la columna 'FECHA' es de tipo datetime en el DataFrame original
    df_diario['FECHA'] = pd.to_datetime(df_diario['FECHA'], format="%d/%m/%Y")

    # Crear una columna para el mes y año
    df_diario['MES'] = df_diario['FECHA'].dt.to_period('M')

    # Calcular los acumulados mensuales
    df_acumulados = df_diario.groupby('MES').sum(numeric_only=True).reset_index()

    # Crear un DataFrame con todos los meses del año
    all_months = pd.date_range(start='2024-01-01', end='2024-12-31', freq='M').to_period('M')
    df_all_months = pd.DataFrame(all_months, columns=['MES'])

    # Unir con el DataFrame de acumulados mensuales para asegurarse de que todos los meses estén presentes
    df_acumulados = pd.merge(df_all_months, df_acumulados, on='MES', how='left')

    # Rellenar los NaN con ceros y asegurarse de que el tipo de datos sea correcto
    df_acumulados.fillna(0, inplace=True)

    # Cambiar el formato del periodo a nombre de mes en español
    df_acumulados['MES'] = df_acumulados['MES'].dt.strftime('%B').str.capitalize()

    # Calcular IDP (Placeholder, ajusta según la fórmula correcta)
    df_IDP = pd.DataFrame({
        'MES': df_acumulados['MES'],
        'DPV': df_acumulados['PREOPERACIONAL'],
        'DTR': df_acumulados['DÍA TRABAJADO'],
        'IDP': (df_acumulados['PREOPERACIONAL'] / df_acumulados['DÍA TRABAJADO']) * 100
    })

    # Redondear a 2 decimales
    df_IDP = df_IDP.round(2)

    # Transponer el DataFrame
    df_IDP = df_IDP.set_index('MES').transpose()

    return df_IDP

def actualizarIndicadores(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac,file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2, file_seguimiento):
   
   # Crear df_dirio y df_hist

    
    df_hist = crearDfHist(file_ubicar1, file_ubicar2, file_mdvr1, file_mdvr2, file_ubicom1, file_ubicom2, file_securitrac, file_wialon1, file_wialon2, file_wialon3, file_ituran1, file_ituran2)
    df_diario = crear_df_diario(df_hist)
    # Calcular los cuatro indicadores

    df_EJL = calcular_EJL(df_diario)
    df_GVE = calcular_GVE(df_diario, df_hist)
    df_ELVL = calcular_ELVL(df_diario)
    df_IDP = calcular_IDP(df_diario)

    # Crear una lista de DataFrames
    dfs = [df_EJL, df_GVE, df_ELVL, df_IDP]
    
    # Crear el archivo si no existe
    if not os.path.exists(file_seguimiento):
        with pd.ExcelWriter(file_seguimiento, engine='xlsxwriter') as writer:
            # Crea un archivo de Excel vacío
            pd.DataFrame().to_excel(writer)
    
    # Cargar el archivo existente
    book = load_workbook(file_seguimiento)
    
    # Crear o seleccionar la hoja de trabajo 'Indicadores'
    if 'Indicadores' in book.sheetnames:
        sheet = book['Indicadores']
    else:
        sheet = book.create_sheet('Indicadores')
    
    # Escribir cada DataFrame en la ubicación especificada con un espacio de 1 fila entre ellos
    start_row = 1
    for df in dfs:
        df = df.reset_index()
        df.rename(columns={'index': 'Indicador'}, inplace=True)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row or c_idx == 1:  # Aplicar formato a los encabezados de los meses y la columna 'Indicador'
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        start_row += len(df) + 2  # Incrementar la fila inicial para el siguiente DataFrame (1 fila de espacio + 1 fila de encabezado)
    
    # Guardar el archivo
    book.save(file_seguimiento)

#### Esto probablemente sea mejor ejectutarlo en otro lado y dejar el script solo con las funciones, pero como esto es un script que en teoría solo se va a usar esta vez vale la pena revisar ###


## ESTAS RUTAS TOCA ACTUALIZARLAS CON LAS MÁS RECIENTES

seguimiento = r"C:\Users\SGI SAS\Desktop\SGI\seguimiento.xlsx"
mdvr_file1 = r"C:\Users\SGI SAS\Downloads\travel_sheet_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720735191.xls"
mdvr_file2 = r"C:\Users\SGI SAS\Downloads\overspeeds_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720734995.xlsx"
ituran_file1 = r"C:\Users\SGI SAS\Downloads\Export.xls"
ituran_file2 = r"C:\Users\SGI SAS\Downloads\Export(22).xls"
securitrac_file = r"C:\Users\SGI SAS\Downloads\exported-excel(5).xls"
wialon_file1 = r"C:\Users\SGI SAS\Downloads\LPN816_INFORME_GENERAL_TM_V1.0_2024-07-11_17-08-35.xlsx"
wialon_file2 = r"C:\Users\SGI SAS\Downloads\LPN821_INFORME_GENERAL_TM_V1.0_2024-07-11_17-10-13.xlsx"
wialon_file3 = r"C:\Users\SGI SAS\Downloads\JTV645_INFORME_GENERAL_TM_V1.0_2024-07-11_17-08-06.xlsx"
ubicar_file1 = r"C:\Users\SGI SAS\Downloads\travel_sheet_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720735468.xls"
ubicar_file2 = r"C:\Users\SGI SAS\Downloads\overspeeds_report_2024_01_01_00_00_00_2024_07_11_00_00_00_1720735429.xlsx"
ubicom_file1 = r"C:\Users\SGI SAS\Downloads\ReporteDiario.xls"
ubicom_file2 = r"C:\Users\SGI SAS\Downloads\Estacionados.xls"


#agregar_datos_historicos(seguimiento, ubicar_file1, ubicar_file2, mdvr_file1, mdvr_file2, ubicom_file1, ubicom_file2, securitrac_file, wialon_file1, wialon_file2, wialon_file3, ituran_file1, ituran_file2)

## Esto llena la hoja de seguimiento en el Excel.

# agregar_datos_historicos(seguimiento, fileUbicar1, fileUbicar2, fileMDVR1, fileMDVR2, fileUbicom1, fileUbicom2, fileSecuritrac, fileWialon1, fileWialon2, fileWialon3, fileIturan1, fileIturan2)

## Esta para que sigue es la que llena la tabla de seguimiento

#df_hist = juntarDatos(fileIturan1, fileIturan2, fileMDVR1, fileMDVR2, fileUbicar1, fileUbicar2, fileUbicom1, fileUbicom2, fileSecuritrac,fileWialon1, fileWialon2, fileWialon3)

# engine = create_engine('mysql+mysqlconnector://root:12345678@localhost:3306/vehiculos')


# # Insertar los datos en la tabla seguimiento
# df_hist.to_sql('seguimiento', con=engine, if_exists='append', index=False)

# print("Datos insertados correctamente en la tabla seguimiento.")

