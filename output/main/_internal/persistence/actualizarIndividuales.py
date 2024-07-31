import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from archivoExcel import extraerIturan, extraerMDVR, extraerSecuritrac, extraerUbicar, extraerUbicom, extraerWialon, infracIturan, infracMDVR, infracSecuritrac, infracUbicar, infracWialon

class ActualizarIndividuales:
    def __init__(self):
        pass

    def actualizar_excel_con_datos(self, output_file, df_nuevos):
        try:
            # Leer Excel 'Seguimiento'.
            book = load_workbook(output_file)
            if 'Seguimiento' in book.sheetnames:
                sheet = book['Seguimiento']
                df_existente = pd.read_excel(output_file, sheet_name='Seguimiento')
            else:
                sheet = book.create_sheet('Seguimiento')
                df_existente = pd.DataFrame()

            # Actualizar el DataFrame con la nueva información.
            for _, row in df_nuevos.iterrows():
                placa = row['placa']
                fecha = pd.to_datetime(row['fecha'], format='%d/%m/%Y')
                dia = fecha.strftime('%d/%m')

                df_existente.loc[(df_existente['PLACA'] == placa) & (df_existente['SEGUIMIENTO'] == 'Nº Excesos'), dia] = row['num_excesos']
                df_existente.loc[(df_existente['PLACA'] == placa) & (df_existente['SEGUIMIENTO'] == 'Nº Desplazamiento'), dia] = row['num_desplazamientos']
                df_existente.loc[(df_existente['PLACA'] == placa) & (df_existente['SEGUIMIENTO'] == 'Día Trabajado'), dia] = row['dia_trabajado']
                df_existente.loc[(df_existente['PLACA'] == placa) & (df_existente['SEGUIMIENTO'] == 'Preoperacional'), dia] = row['preoperacional']
                df_existente.loc[(df_existente['PLACA'] == placa) & (df_existente['SEGUIMIENTO'] == 'Km recorridos'), dia] = row['km_recorridos']

            # Llenar el Excel.
            for row_idx, row in enumerate(dataframe_to_rows(df_existente, index=False, header=True), 1):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            book.save(output_file)
        except Exception as e:
            print(f"Error al actualizar el archivo Excel: {e}")

    def actualizar_excel_con_infracciones(self, output_file, df_infracciones):
        try:
            # Leer el archivo Excel existente
            book = load_workbook(output_file)
            if 'Infracciones' in book.sheetnames:
                sheet = book['Infracciones']
                df_existente = pd.read_excel(output_file, sheet_name='Infracciones')
            else:
                sheet = book.create_sheet('Infracciones')
                df_existente = pd.DataFrame()

            # Concatenar los datos nuevos con los existentes
            df_total = pd.concat([df_existente, df_infracciones], ignore_index=True)

            # Escribir el DataFrame en la hoja de Excel
            for row_idx, row in enumerate(dataframe_to_rows(df_total, index=False, header=True), 1):
                for col_idx, value in enumerate(row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            book.save(output_file)
        except Exception as e:
            print(f"Error al actualizar el archivo de infracciones: {e}")

    def llenarInfracIturan(self, file_Ituran, output_file):
        try:
            infracciones = infracIturan(file_Ituran)
            df_infracciones = pd.DataFrame(infracciones)
            df_infracciones['FECHA'] = pd.to_datetime(df_infracciones['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')
            self.actualizar_excel_con_infracciones(output_file, df_infracciones)
        except Exception as e:
            print(f"Error al llenar infracciones de Ituran: {e}")

    def llenarInfracMDVR(self, file_MDVR, output_file):
        try:
            infracciones = infracMDVR(file_MDVR)
            df_infracciones = pd.DataFrame(infracciones)
            df_infracciones['FECHA'] = pd.to_datetime(df_infracciones['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')
            self.actualizar_excel_con_infracciones(output_file, df_infracciones)
        except Exception as e:
            print(f"Error al llenar infracciones de MDVR: {e}")

    def llenarInfracUbicar(self, file_Ubicar, output_file):
        try:
            infracciones = infracUbicar(file_Ubicar)
            df_infracciones = pd.DataFrame(infracciones)
            df_infracciones['FECHA'] = pd.to_datetime(df_infracciones['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')
            self.actualizar_excel_con_infracciones(output_file, df_infracciones)
        except Exception as e:
            print(f"Error al llenar infracciones de Ubicar: {e}")

    def llenarInfracSecuritrac(self, file_Securitrac, output_file):
        try:
            infracciones = infracSecuritrac(file_Securitrac)
            df_infracciones = pd.DataFrame(infracciones)
            df_infracciones['FECHA'] = pd.to_datetime(df_infracciones['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')
            self.actualizar_excel_con_infracciones(output_file, df_infracciones)
        except Exception as e:
            print(f"Error al llenar infracciones de Securitrac: {e}")

    def llenarInfracWialon(self, file_Wialon1, file_Wialon2, file_Wialon3, output_file):
        try:
            infrac1 = infracWialon(file_Wialon1)
            infrac2 = infracWialon(file_Wialon2)
            infrac3 = infracWialon(file_Wialon3)
            infracciones = infrac1 + infrac2 + infrac3
            df_infracciones = pd.DataFrame(infracciones)
            df_infracciones['FECHA'] = pd.to_datetime(df_infracciones['FECHA'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y %H:%M:%S')
            self.actualizar_excel_con_infracciones(output_file, df_infracciones)
        except Exception as e:
            print(f"Error al llenar infracciones de Securitrac: {e}")   


    def llenarIturan(self, ituran_file1, ituran_file2, output_file):
        try:
            nuevos_datos = extraerIturan(ituran_file1, ituran_file2)
            df_nuevos = pd.DataFrame(nuevos_datos)
            self.actualizar_excel_con_datos(output_file, df_nuevos)
        except Exception as e:
            print(f"Error al llenar datos de Ituran: {e}")

    def llenarMDVR(self, mdvr_file1, mdvr_file2, output_file):
        try:
            nuevos_datos = extraerMDVR(mdvr_file1, mdvr_file2)
            df_nuevos = pd.DataFrame(nuevos_datos)
            self.actualizar_excel_con_datos(output_file, df_nuevos)
        except Exception as e:
            print(f"Error al llenar datos de MDVR: {e}")

    def llenarUbicar(self, ubicar_file1, ubicar_file2, output_file):
        try:
            nuevos_datos = extraerUbicar(ubicar_file1, ubicar_file2)
            df_nuevos = pd.DataFrame(nuevos_datos)
            self.actualizar_excel_con_datos(output_file, df_nuevos)
        except Exception as e:
            print(f"Error al llenar datos de Ubicar: {e}")

    def llenarUbicom(self, ubicom_file1, ubicom_file2, output_file):
        try:
            nuevos_datos = extraerUbicom(ubicom_file1, ubicom_file2)
            df_nuevos = pd.DataFrame(nuevos_datos)
            self.actualizar_excel_con_datos(output_file, df_nuevos)
        except Exception as e:
            print(f"Error al llenar datos de Ubicom: {e}")

    def llenarWialon(self, wialon_file1, wialon_file2, wialon_file3, output_file):
        try:
            nuevos_datos = extraerWialon(wialon_file1, wialon_file2, wialon_file3)
            df_nuevos = pd.DataFrame(nuevos_datos)
            self.actualizar_excel_con_datos(output_file, df_nuevos)
        except Exception as e:
            print(f"Error al llenar datos de Wialon: {e}")

    def llenarSecuritrac(self, securitrac_file, output_file):
        try:
            nuevos_datos = extraerSecuritrac(securitrac_file)
            df_nuevos = pd.DataFrame(nuevos_datos)
            self.actualizar_excel_con_datos(output_file, df_nuevos)
        except Exception as e:
            print(f"Error al llenar datos de Securitrac: {e}")